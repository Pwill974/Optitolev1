import csv
import io
import math
import random
import re
import tempfile
import time
import unicodedata
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from numbers import Number
from pathlib import Path
from typing import Iterable

import ezdxf
from ezdxf import edgeminer, edgesmith
import matplotlib.pyplot as plt
import pandas as pd
import streamlit as st
from ezdxf.path import from_hatch, make_path
from openpyxl import load_workbook
from shapely import affinity
from shapely.geometry import LinearRing, Polygon, box
from shapely.ops import unary_union
from shapely.validation import make_valid


# ============================================================
# Modèles de données
# ============================================================

@dataclass
class NomenclatureItem:
    reference_display: str
    reference_key: str
    quantity: int
    thickness: str
    material: str


@dataclass
class DxfPiece:
    reference_display: str
    reference_key: str
    source_name: str
    polygon: Polygon
    quantity: int
    thickness: str
    material: str


@dataclass
class Placement:
    reference: str
    source_name: str
    sheet_index: int
    rotation: int
    polygon: Polygon
    original_polygon: Polygon
    copy_index: int
    thickness: str
    material: str


# ============================================================
# Normalisation
# ============================================================

REFERENCE_HEADERS = {
    "repere", "reperepiece", "numerorepere", "numerodepiece",
    "numeropiece", "piece", "position", "mark", "partmark",
    "mainpartmark",
}
QUANTITY_HEADERS = {
    "quantite", "qte", "nombre", "nb", "quantity", "qty",
}
THICKNESS_HEADERS = {
    "epaisseur", "ep", "thickness", "plate thickness", "thk",
}
MATERIAL_HEADERS = {
    "matiere", "materiau", "material", "nuance", "grade", "steelgrade",
}


def remove_accents(text: str) -> str:
    return "".join(
        char for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, Number):
        number = float(value)
        if number.is_integer():
            return str(int(number))
    return str(value).strip()


def normalize_reference(value: object) -> str:
    text = remove_accents(cell_to_text(value)).upper()
    text = re.sub(r"[\s\-_]+", "", text)
    if not text:
        return ""
    if text.isdigit():
        return str(int(text))
    return text


def normalize_header(value: object) -> str:
    text = remove_accents(cell_to_text(value)).lower()
    return re.sub(r"[^a-z0-9]+", "", text)


def reference_from_dxf_filename(filename: str) -> str:
    """
    Extrait le repère réel depuis un nom de fichier Advance Steel.

    Exemples :
        NC/13.nc.dxf          -> 13
        NC/AT1.nc.dxf         -> AT1
        NC/AT1.nc.err.dxf     -> AT1
        NC/AT1A11.nc.err.dxf  -> AT1A11
        10.dxf                -> 10
    """
    filename_only = Path(filename).name

    # Retirer l'extension finale .dxf.
    cleaned = re.sub(r"(?i)\.dxf$", "", filename_only).strip()

    # Retirer successivement les suffixes techniques ajoutés par Advance Steel.
    technical_suffixes = ("nc", "err", "dstv", "cnc", "cam")

    while True:
        previous = cleaned
        suffix_pattern = r"(?i)\.(?:" + "|".join(technical_suffixes) + r")$"
        cleaned = re.sub(suffix_pattern, "", cleaned).strip()

        if cleaned == previous:
            break

    return normalize_reference(cleaned)


def normalize_group_value(value: object, fallback: str) -> str:
    text = cell_to_text(value)
    return text if text else fallback


# ============================================================
# Lecture de la nomenclature
# ============================================================

def read_excel_rows(data: bytes) -> list[list[object]]:
    """
    Recherche automatiquement une feuille Excel contenant les colonnes
    Repère et Quantité. Cela évite l'erreur lorsque la feuille active du
    classeur est absente, masquée ou invalide.
    """
    try:
        workbook = load_workbook(
            io.BytesIO(data),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ValueError(
            "Impossible d'ouvrir le fichier Excel. "
            "Vérifiez qu'il s'agit bien d'un fichier .xlsx valide."
        ) from exc

    worksheets = list(workbook.worksheets)

    if not worksheets:
        workbook.close()
        raise ValueError(
            "Le fichier Excel ne contient aucune feuille de calcul exploitable."
        )

    first_non_empty_rows = None

    try:
        for worksheet in worksheets:
            rows = [
                list(row)
                for row in worksheet.iter_rows(values_only=True)
            ]

            if not rows or not any(
                any(cell not in (None, "") for cell in row)
                for row in rows
            ):
                continue

            if first_non_empty_rows is None:
                first_non_empty_rows = rows

            try:
                find_header_row(rows)
                return rows
            except ValueError:
                continue
    finally:
        workbook.close()

    if first_non_empty_rows is not None:
        return first_non_empty_rows

    raise ValueError(
        "Toutes les feuilles du fichier Excel sont vides."
    )


def read_csv_rows(data: bytes) -> list[list[object]]:
    decoded_text = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            decoded_text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if decoded_text is None:
        raise ValueError("Impossible de lire le fichier CSV.")

    try:
        dialect = csv.Sniffer().sniff(decoded_text[:4096], delimiters=";,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    return [list(row) for row in csv.reader(io.StringIO(decoded_text), dialect)]


def find_column(normalized_headers: list[str], accepted: set[str]) -> int | None:
    return next(
        (index for index, value in enumerate(normalized_headers) if value in accepted),
        None,
    )


def find_header_row(rows: list[list[object]]) -> tuple[int, dict[str, int | None]]:
    for row_index, row in enumerate(rows[:40]):
        normalized = [normalize_header(value) for value in row]
        reference_index = find_column(normalized, REFERENCE_HEADERS)
        quantity_index = find_column(normalized, QUANTITY_HEADERS)

        if reference_index is not None and quantity_index is not None:
            return row_index, {
                "reference": reference_index,
                "quantity": quantity_index,
                "thickness": find_column(normalized, THICKNESS_HEADERS),
                "material": find_column(normalized, MATERIAL_HEADERS),
            }

    raise ValueError(
        "Les colonnes Repère et Quantité sont introuvables. "
        "Utilisez au minimum les titres « Repère » et « Quantité »."
    )


def parse_quantity(value: object, row_number: int) -> int:
    if value is None or str(value).strip() == "":
        raise ValueError(f"Quantité vide à la ligne {row_number}.")
    try:
        quantity = int(float(str(value).replace(",", ".").strip()))
    except ValueError as exc:
        raise ValueError(
            f"Quantité invalide à la ligne {row_number} : {value!r}"
        ) from exc
    if quantity <= 0:
        raise ValueError(
            f"La quantité doit être supérieure à zéro à la ligne {row_number}."
        )
    return quantity


def read_nomenclature(uploaded_file) -> list[NomenclatureItem]:
    data = uploaded_file.getvalue()
    suffix = Path(uploaded_file.name).suffix.lower()

    if suffix == ".xlsx":
        rows = read_excel_rows(data)
    elif suffix == ".csv":
        rows = read_csv_rows(data)
    else:
        raise ValueError("Utilisez une nomenclature au format .xlsx ou .csv.")

    header_row, columns = find_header_row(rows)
    result: list[NomenclatureItem] = []

    for row_number, row in enumerate(rows[header_row + 1:], start=header_row + 2):
        def get_cell(index: int | None):
            if index is None or index >= len(row):
                return None
            return row[index]

        reference_value = get_cell(columns["reference"])
        reference_display = cell_to_text(reference_value)
        reference_key = normalize_reference(reference_value)

        if not reference_key:
            continue

        result.append(
            NomenclatureItem(
                reference_display=reference_display,
                reference_key=reference_key,
                quantity=parse_quantity(get_cell(columns["quantity"]), row_number),
                thickness=normalize_group_value(
                    get_cell(columns["thickness"]), "Non renseignée"
                ),
                material=normalize_group_value(
                    get_cell(columns["material"]), "Non renseignée"
                ),
            )
        )

    if not result:
        raise ValueError("Aucune pièce exploitable n'a été trouvée.")

    return result


# ============================================================
# Lecture des DXF
# ============================================================

def flatten_path(path, tolerance: float) -> list[tuple[float, float]]:
    try:
        return [
            (float(vertex.x), float(vertex.y))
            for vertex in path.flattening(max(tolerance, 0.05))
        ]
    except Exception:
        return []


def flatten_entity(entity, tolerance: float) -> list[tuple[float, float]]:
    try:
        return flatten_path(make_path(entity), tolerance)
    except Exception:
        return []


def clean_ring(points: list[tuple[float, float]]) -> list[tuple[float, float]]:
    cleaned: list[tuple[float, float]] = []

    for x, y in points:
        point = (round(float(x), 6), round(float(y), 6))
        if not cleaned or point != cleaned[-1]:
            cleaned.append(point)

    if len(cleaned) >= 2 and cleaned[0] == cleaned[-1]:
        cleaned.pop()

    return cleaned


def entity_is_closed(entity) -> bool:
    """
    Détecte les contours réellement fermés, même lorsque le drapeau DXF
    « closed » est faux.

    Advance Steel exporte notamment certains trous circulaires sous forme
    de LWPOLYLINE annoncées ouvertes, mais composées de deux arcs avec un
    point final identique au point initial. ezdxf reconnaît alors le chemin
    comme fermé : il faut donc tester la géométrie et pas seulement le drapeau.
    """
    entity_type = entity.dxftype()

    if entity_type == "LWPOLYLINE":
        if bool(entity.closed):
            return True

        try:
            path = make_path(entity)
            if path.is_closed:
                return True
        except Exception:
            pass

        try:
            points = list(entity.get_points("xy"))
            if len(points) >= 3:
                first_x, first_y = points[0]
                last_x, last_y = points[-1]
                return math.hypot(
                    float(first_x) - float(last_x),
                    float(first_y) - float(last_y),
                ) <= 0.05
        except Exception:
            return False

        return False

    if entity_type == "POLYLINE":
        if bool(entity.is_closed):
            return True

        try:
            path = make_path(entity)
            return bool(path.is_closed)
        except Exception:
            return False

    if entity_type in {"CIRCLE", "ELLIPSE"}:
        return True

    if entity_type == "SPLINE":
        if bool(getattr(entity, "closed", False)):
            return True

        try:
            return bool(make_path(entity).is_closed)
        except Exception:
            return False

    return False


def polygon_from_ring_points(
    points: list[tuple[float, float]],
    minimum_area: float = 0.01,
) -> Polygon | None:
    points = clean_ring(points)

    if len(points) < 3:
        return None

    try:
        polygon = Polygon(LinearRing(points))
    except Exception:
        return None

    if polygon.is_empty or polygon.area <= minimum_area:
        return None

    if not polygon.is_valid:
        polygon = make_valid(polygon)

    if polygon.geom_type == "MultiPolygon":
        polygon = max(polygon.geoms, key=lambda part: part.area)

    if polygon.geom_type != "Polygon" or polygon.area <= minimum_area:
        return None

    return polygon


def deduplicate_polygons(polygons: list[Polygon]) -> list[Polygon]:
    result: list[Polygon] = []

    for polygon in sorted(polygons, key=lambda item: item.area, reverse=True):
        duplicate = False

        for existing in result:
            area_scale = max(existing.area, polygon.area, 1.0)

            if abs(existing.area - polygon.area) / area_scale > 1e-5:
                continue

            if existing.hausdorff_distance(polygon) <= 0.01:
                duplicate = True
                break

        if not duplicate:
            result.append(polygon)

    return result


def iter_geometry_entities(layout, maximum_depth: int = 8):
    """
    Parcourt les entités du modelspace et développe récursivement les INSERT.

    Advance Steel peut stocker certains contours ou perçages dans des blocs.
    INSERT.virtual_entities() renvoie les entités déjà transformées en coordonnées
    du dessin, sans modifier le fichier source.
    """
    stack = [(entity, 0) for entity in layout]

    while stack:
        entity, depth = stack.pop(0)

        if entity.dxftype() == "INSERT" and depth < maximum_depth:
            try:
                virtual = list(entity.virtual_entities())
                stack[0:0] = [(child, depth + 1) for child in virtual]
            except Exception:
                continue
        else:
            yield entity


def extract_direct_closed_loops(entities, tolerance: float) -> list[Polygon]:
    polygons: list[Polygon] = []

    for entity in entities:
        entity_type = entity.dxftype()

        if entity_type in {"HATCH", "MPOLYGON"}:
            try:
                for path in from_hatch(entity):
                    polygon = polygon_from_ring_points(
                        flatten_path(path, tolerance)
                    )
                    if polygon is not None:
                        polygons.append(polygon)
            except Exception:
                continue

        elif entity_is_closed(entity):
            polygon = polygon_from_ring_points(
                flatten_entity(entity, tolerance)
            )
            if polygon is not None:
                polygons.append(polygon)

    return polygons


def extract_connected_edge_loops(
    entities,
    tolerance: float,
    gap_tolerance: float,
) -> list[Polygon]:
    """Reconstruit les contours composés de LINE, ARC et courbes ouvertes."""
    polygons: list[Polygon] = []

    try:
        open_entities = list(edgesmith.filter_open_edges(entities))
        edges = list(
            edgesmith.edges_from_entities_2d(
                open_entities,
                gap_tol=gap_tolerance,
            )
        )

        if len(edges) < 2:
            return polygons

        deposit = edgeminer.Deposit(edges, gap_tol=gap_tolerance)
        loops = edgeminer.find_all_loops(deposit, timeout=8.0)

        for loop in loops:
            try:
                path = edgesmith.path2d_from_chain(loop)
                polygon = polygon_from_ring_points(
                    flatten_path(path, tolerance)
                )
                if polygon is not None:
                    polygons.append(polygon)
            except Exception:
                continue

    except Exception:
        # Les contours fermés directs restent exploitables même si un DXF
        # contient un réseau de lignes trop complexe pour EdgeMiner.
        return polygons

    return polygons


def build_plate_polygon(loop_polygons: list[Polygon], source_name: str) -> Polygon:
    loops = deduplicate_polygons(loop_polygons)

    if not loops:
        raise ValueError(
            f"Aucun contour fermé exploitable dans {source_name}."
        )

    loops.sort(key=lambda polygon: polygon.area, reverse=True)
    outer = loops[0]
    inside_loops = []

    for candidate in loops[1:]:
        point = candidate.representative_point()
        if outer.covers(point):
            inside_loops.append(candidate)

    # Les trous directs sont les boucles dont le plus petit contenant est
    # le contour extérieur. Les éventuels îlots imbriqués ne deviennent pas
    # de faux trous supplémentaires.
    direct_holes: list[Polygon] = []

    for candidate in inside_loops:
        candidate_point = candidate.representative_point()
        has_inner_parent = any(
            other.area > candidate.area
            and other.covers(candidate_point)
            for other in inside_loops
            if other is not candidate
        )

        if not has_inner_parent:
            direct_holes.append(candidate)

    plate = Polygon(
        outer.exterior.coords[:],
        [hole.exterior.coords[:] for hole in direct_holes],
    )

    if not plate.is_valid:
        plate = make_valid(plate)

    if plate.geom_type == "MultiPolygon":
        plate = max(plate.geoms, key=lambda part: part.area)

    if plate.geom_type != "Polygon" or plate.is_empty:
        raise ValueError(
            f"La géométrie de {source_name} est invalide après reconstruction."
        )

    min_x, min_y, _, _ = plate.bounds
    return affinity.translate(plate, xoff=-min_x, yoff=-min_y)


def polygons_from_dxf_bytes(
    data: bytes,
    source_name: str,
    tolerance: float,
) -> list[Polygon]:
    with tempfile.NamedTemporaryFile(suffix=".dxf", delete=False) as tmp:
        tmp.write(data)
        temp_name = tmp.name

    try:
        document = ezdxf.readfile(temp_name)
        modelspace = document.modelspace()
        geometry_entities = list(iter_geometry_entities(modelspace))

        gap_tolerance = max(0.02, min(0.5, tolerance * 0.2))

        loop_polygons = extract_direct_closed_loops(
            geometry_entities,
            tolerance,
        )
        loop_polygons.extend(
            extract_connected_edge_loops(
                geometry_entities,
                tolerance,
                gap_tolerance,
            )
        )

        return [build_plate_polygon(loop_polygons, source_name)]

    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"DXF illisible : {source_name}") from exc
    finally:
        try:
            Path(temp_name).unlink(missing_ok=True)
        except Exception:
            pass


def read_dxf_zip(uploaded_file) -> dict[str, tuple[str, bytes]]:
    data = uploaded_file.getvalue()

    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            result: dict[str, tuple[str, bytes]] = {}
            for name in archive.namelist():
                if name.endswith("/") or Path(name).suffix.lower() != ".dxf":
                    continue

                key = reference_from_dxf_filename(name)
                if not key:
                    continue

                if key in result:
                    previous_name = result[key][0]
                    raise ValueError(
                        f"Deux DXF correspondent au même repère : "
                        f"{previous_name} et {name}."
                    )

                result[key] = (name, archive.read(name))
    except zipfile.BadZipFile as exc:
        raise ValueError("Le fichier ZIP des DXF est invalide.") from exc

    if not result:
        raise ValueError("Aucun fichier DXF n'a été trouvé dans le ZIP.")

    return result


def build_pieces(
    nomenclature: list[NomenclatureItem],
    dxf_files: dict[str, tuple[str, bytes]],
    tolerance: float,
) -> tuple[list[DxfPiece], list[str], list[str]]:
    pieces: list[DxfPiece] = []
    missing: list[str] = []
    used_keys: set[str] = set()

    for item in nomenclature:
        dxf_info = dxf_files.get(item.reference_key)
        if dxf_info is None:
            missing.append(item.reference_display)
            continue

        source_name, data = dxf_info
        polygons = polygons_from_dxf_bytes(data, source_name, tolerance)
        polygon = polygons[0]

        pieces.append(
            DxfPiece(
                reference_display=item.reference_display,
                reference_key=item.reference_key,
                source_name=source_name,
                polygon=polygon,
                quantity=item.quantity,
                thickness=item.thickness,
                material=item.material,
            )
        )
        used_keys.add(item.reference_key)

    unused = [
        source_name for key, (source_name, _) in dxf_files.items()
        if key not in used_keys
    ]

    return pieces, missing, unused


# ============================================================
# Moteur de nesting heuristique
# ============================================================


# ============================================================
# V14 — Rotations naturelles et moteur dense
# ============================================================

USE_NATURAL_EDGE_ANGLES = True
MAX_NATURAL_ROTATIONS = 36
DENSE_CANDIDATE_MULTIPLIER = 1.0
FILL_PRIORITY = "Hauteur 1500 mm"


def normalize_angle(angle: float) -> float:
    value = float(angle) % 360.0
    if abs(value - round(value)) < 1e-6:
        return float(int(round(value)) % 360)
    return round(value, 3)


def unique_angles(angles: list[float], tolerance: float = 0.25) -> list[float]:
    result: list[float] = []

    for angle in angles:
        candidate = normalize_angle(angle)
        if not any(
            abs(((candidate - existing + 180.0) % 360.0) - 180.0) <= tolerance
            for existing in result
        ):
            result.append(candidate)

    return result


def dominant_edge_angles(polygon: Polygon, limit: int = 8) -> list[float]:
    segments = []
    coords = list(polygon.exterior.coords)

    for index in range(len(coords) - 1):
        x1, y1 = coords[index]
        x2, y2 = coords[index + 1]
        length = math.hypot(x2 - x1, y2 - y1)

        if length < 10.0:
            continue

        angle = math.degrees(math.atan2(y2 - y1, x2 - x1)) % 180.0
        segments.append((length, round(angle, 3)))

    segments.sort(reverse=True)
    angles = []

    for _, angle in segments:
        if not any(abs(angle - existing) <= 1.0 for existing in angles):
            angles.append(angle)
        if len(angles) >= limit:
            break

    return angles


def piece_rotation_angles(piece: DxfPiece, base_rotations: list[float]) -> list[float]:
    angles = [float(angle) for angle in base_rotations]

    if USE_NATURAL_EDGE_ANGLES:
        for edge_angle in dominant_edge_angles(piece.polygon):
            angles.extend(
                [
                    -edge_angle,
                    90.0 - edge_angle,
                    180.0 - edge_angle,
                    270.0 - edge_angle,
                ]
            )

    angles = unique_angles(angles)

    if len(angles) > MAX_NATURAL_ROTATIONS:
        base = unique_angles([float(angle) for angle in base_rotations])
        extra = [angle for angle in angles if angle not in base]
        angles = base + extra[: max(0, MAX_NATURAL_ROTATIONS - len(base))]

    return angles


def piece_like_from_placement(item: Placement) -> DxfPiece:
    return DxfPiece(
        reference_display=item.reference,
        reference_key=normalize_reference(item.reference),
        source_name=item.source_name,
        polygon=item.original_polygon,
        quantity=1,
        thickness=item.thickness,
        material=item.material,
    )


def is_height_priority() -> bool:
    return str(FILL_PRIORITY).startswith("Hauteur")


def is_width_priority() -> bool:
    return str(FILL_PRIORITY).startswith("Largeur")



def rotated_at_origin(polygon: Polygon, angle: int) -> Polygon:
    rotated = affinity.rotate(
        polygon,
        angle,
        origin=(0, 0),
        use_radians=False,
    )
    min_x, min_y, _, _ = rotated.bounds
    return affinity.translate(rotated, xoff=-min_x, yoff=-min_y)


def sampled_coordinates(ring, maximum: int = 12) -> list[tuple[float, float]]:
    coordinates = list(ring.coords)

    if len(coordinates) <= maximum:
        return [(float(x), float(y)) for x, y in coordinates]

    step = max(1, len(coordinates) // maximum)
    return [
        (float(coordinates[index][0]), float(coordinates[index][1]))
        for index in range(0, len(coordinates), step)
    ][:maximum]


def sampled_segments(ring, maximum: int = 10):
    coordinates = list(ring.coords)
    segments = []

    for index in range(len(coordinates) - 1):
        x1, y1 = coordinates[index]
        x2, y2 = coordinates[index + 1]
        length = math.hypot(x2 - x1, y2 - y1)
        if length > 0.5:
            segments.append(
                (
                    (float(x1), float(y1)),
                    (float(x2), float(y2)),
                    float(length),
                )
            )

    if len(segments) <= maximum:
        return segments

    step = max(1, len(segments) // maximum)
    return segments[::step][:maximum]


def candidate_positions(
    rotated_piece: Polygon,
    placed_polygons: list[Polygon],
    margin: float,
    clearance: float,
    sheet_width: float,
    sheet_height: float,
    allow_hole_nesting: bool,
    quality_level: int,
    max_candidates: int,
) -> list[tuple[float, float]]:
    """
    Génère des positions true-shape approchées : bords, sommets et segments
    parallèles. L'alignement segment contre segment améliore fortement
    l'emboîtement des goussets et platines inclinées.
    """
    width = rotated_piece.bounds[2] - rotated_piece.bounds[0]
    height = rotated_piece.bounds[3] - rotated_piece.bounds[1]
    candidates = {
        (round(margin, 4), round(margin, 4)),
        (round(sheet_width - margin - width, 4), round(margin, 4)),
        (round(margin, 4), round(sheet_height - margin - height, 4)),
        (
            round(sheet_width - margin - width, 4),
            round(sheet_height - margin - height, 4),
        ),
    }

    moving_vertices = sampled_coordinates(
        rotated_piece.exterior,
        maximum=8 if quality_level <= 1 else 16,
    )
    moving_segments = sampled_segments(
        rotated_piece.exterior,
        maximum=6 if quality_level <= 1 else 12,
    )

    def add_candidate(x: float, y: float) -> None:
        if (
            x >= margin - 1e-6
            and y >= margin - 1e-6
            and x + width <= sheet_width - margin + 1e-6
            and y + height <= sheet_height - margin + 1e-6
        ):
            candidates.add((round(x, 4), round(y, 4)))

    for polygon in placed_polygons:
        min_x, min_y, max_x, max_y = polygon.bounds

        for x, y in (
            (max_x + clearance, min_y),
            (max_x + clearance, margin),
            (min_x, max_y + clearance),
            (margin, max_y + clearance),
            (max_x + clearance, max_y + clearance),

            (min_x - clearance - width, min_y),
            (min_x - clearance - width, margin),
            (min_x, min_y - clearance - height),
            (margin, min_y - clearance - height),

            (max_x + clearance, max_y - height),
            (min_x - clearance - width, max_y - height),
            (max_x - width, max_y + clearance),
            (max_x - width, min_y - clearance - height),

            (sheet_width - margin - width, min_y),
            (min_x, sheet_height - margin - height),

            # Spécial remplissage 1500 mm : empilement vertical strict
            # dans la même colonne avant ouverture de la colonne suivante.
            (min_x, max_y + clearance),
            (min_x + clearance, max_y + clearance),
            (max_x - width, max_y + clearance),
            (max_x - width, margin),
            (min_x, sheet_height - margin - height),
            (max_x + clearance, sheet_height - margin - height),
        ):
            add_candidate(x, y)

        fixed_vertices = sampled_coordinates(
            polygon.exterior,
            maximum=8 if quality_level <= 1 else 18,
        )

        # Alignement sommet-sommet.
        if quality_level >= 1:
            for fixed_x, fixed_y in fixed_vertices:
                for moving_x, moving_y in moving_vertices:
                    add_candidate(
                        fixed_x + clearance - moving_x,
                        fixed_y - moving_y,
                    )
                    add_candidate(
                        fixed_x - clearance - moving_x,
                        fixed_y - moving_y,
                    )
                    add_candidate(
                        fixed_x - moving_x,
                        fixed_y + clearance - moving_y,
                    )
                    add_candidate(
                        fixed_x - moving_x,
                        fixed_y - clearance - moving_y,
                    )
                    if len(candidates) >= max_candidates * 4:
                        break
                if len(candidates) >= max_candidates * 4:
                    break

        # Alignement bord contre bord pour emboîter les formes inclinées.
        if quality_level >= 2 and len(candidates) < max_candidates * 4:
            fixed_segments = sampled_segments(
                polygon.exterior,
                maximum=8 if quality_level == 2 else 14,
            )

            for (fa, fb, fixed_length) in fixed_segments:
                fdx = fb[0] - fa[0]
                fdy = fb[1] - fa[1]
                fux = fdx / fixed_length
                fuy = fdy / fixed_length
                normal_x = -fuy
                normal_y = fux
                fixed_mid = ((fa[0] + fb[0]) / 2.0, (fa[1] + fb[1]) / 2.0)

                for (ma, mb, moving_length) in moving_segments:
                    mdx = mb[0] - ma[0]
                    mdy = mb[1] - ma[1]
                    mux = mdx / moving_length
                    muy = mdy / moving_length
                    parallel_error = abs(fux * muy - fuy * mux)

                    if parallel_error > 0.16:
                        continue

                    moving_mid = (
                        (ma[0] + mb[0]) / 2.0,
                        (ma[1] + mb[1]) / 2.0,
                    )
                    anchor_pairs = (
                        (fa, mb),
                        (fb, ma),
                        (fixed_mid, moving_mid),
                    )

                    for fixed_anchor, moving_anchor in anchor_pairs:
                        for sign in (-1.0, 1.0):
                            add_candidate(
                                fixed_anchor[0]
                                - moving_anchor[0]
                                + sign * normal_x * clearance,
                                fixed_anchor[1]
                                - moving_anchor[1]
                                + sign * normal_y * clearance,
                            )

                    if len(candidates) >= max_candidates * 4:
                        break
                if len(candidates) >= max_candidates * 4:
                    break

        if allow_hole_nesting:
            for interior in polygon.interiors:
                hole = Polygon(interior)
                hole_min_x, hole_min_y, hole_max_x, hole_max_y = hole.bounds
                hole_width = hole_max_x - hole_min_x
                hole_height = hole_max_y - hole_min_y

                if (
                    hole_width + 1e-6 >= width + 2 * clearance
                    and hole_height + 1e-6 >= height + 2 * clearance
                ):
                    for x, y in (
                        (hole_min_x + clearance, hole_min_y + clearance),
                        (hole_max_x - clearance - width, hole_min_y + clearance),
                        (hole_min_x + clearance, hole_max_y - clearance - height),
                        (hole.centroid.x - width / 2.0, hole.centroid.y - height / 2.0),
                    ):
                        add_candidate(x, y)

        if len(candidates) >= max_candidates * 4:
            break

    # Ordre des positions selon le sens de remplissage choisi.
    # Hauteur 1500 mm : on remplit d'abord en vertical,
    # puis on avance en largeur.
    if is_height_priority():
        ordered = sorted(
            candidates,
            key=lambda position: (
                position[0] * sheet_height + position[1],
                position[0],
                position[1],
            ),
        )
    elif is_width_priority():
        ordered = sorted(
            candidates,
            key=lambda position: (
                position[1] * sheet_width + position[0],
                position[1],
                position[0],
            ),
        )
    else:
        ordered = sorted(
            candidates,
            key=lambda position: (
                position[1] * sheet_width + position[0],
                position[0] * sheet_height + position[1],
            ),
        )

    max_candidates = max(20, int(max_candidates * DENSE_CANDIDATE_MULTIPLIER))

    if len(ordered) <= max_candidates:
        return ordered

    primary_count = max(1, int(max_candidates * 0.70))
    primary = ordered[:primary_count]
    remaining = ordered[primary_count:]
    extra_count = max_candidates - primary_count

    if extra_count > 0 and remaining:
        step = max(1, len(remaining) // extra_count)
        primary.extend(remaining[::step][:extra_count])

    return primary[:max_candidates]


def bounds_are_close(
    first_bounds: tuple[float, float, float, float],
    second_bounds: tuple[float, float, float, float],
    clearance: float,
) -> bool:
    first_min_x, first_min_y, first_max_x, first_max_y = first_bounds
    second_min_x, second_min_y, second_max_x, second_max_y = second_bounds

    if first_max_x + clearance <= second_min_x:
        return False
    if second_max_x + clearance <= first_min_x:
        return False
    if first_max_y + clearance <= second_min_y:
        return False
    if second_max_y + clearance <= first_min_y:
        return False

    return True


def fits_on_sheet(
    polygon: Polygon,
    placed_polygons: list[Polygon],
    sheet_inner: Polygon,
    clearance: float,
) -> bool:
    if not sheet_inner.covers(polygon):
        return False

    candidate_bounds = polygon.bounds

    for other in placed_polygons:
        if not bounds_are_close(
            candidate_bounds,
            other.bounds,
            clearance,
        ):
            continue

        if polygon.intersects(other):
            return False

        if clearance > 0 and polygon.distance(other) < clearance - 1e-7:
            return False

    return True


def directional_compact(
    polygon: Polygon,
    placed_polygons: list[Polygon],
    sheet_inner: Polygon,
    clearance: float,
    minimum_step: float = 0.5,
) -> Polygon:
    """Compacte vers la gauche, le bas et les diagonales bas-gauche."""
    result = polygon
    directions = (
        (-1.0, 0.0),
        (0.0, -1.0),
        (-1.0, -1.0),
        (-1.0, 0.0),
        (0.0, -1.0),
    )

    for dx, dy in directions:
        min_x, min_y, max_x, max_y = result.bounds
        limits = []

        if dx < 0:
            limits.append(min_x - sheet_inner.bounds[0])
        elif dx > 0:
            limits.append(sheet_inner.bounds[2] - max_x)

        if dy < 0:
            limits.append(min_y - sheet_inner.bounds[1])
        elif dy > 0:
            limits.append(sheet_inner.bounds[3] - max_y)

        if not limits:
            continue

        step = max(0.0, min(limits))

        while step >= minimum_step:
            moved = affinity.translate(
                result,
                xoff=dx * step,
                yoff=dy * step,
            )

            if fits_on_sheet(
                moved,
                placed_polygons,
                sheet_inner,
                clearance,
            ):
                result = moved
                min_x, min_y, max_x, max_y = result.bounds
                new_limits = []
                if dx < 0:
                    new_limits.append(min_x - sheet_inner.bounds[0])
                if dy < 0:
                    new_limits.append(min_y - sheet_inner.bounds[1])
                step = max(0.0, min(new_limits)) if new_limits else 0.0
            else:
                step /= 2.0

    return result


def contact_count(
    polygon: Polygon,
    placed_polygons: list[Polygon],
    sheet_inner: Polygon,
    clearance: float,
) -> int:
    tolerance = max(0.75, clearance * 0.15)
    contacts = 0
    min_x, min_y, max_x, max_y = polygon.bounds
    inner_min_x, inner_min_y, inner_max_x, inner_max_y = sheet_inner.bounds

    if abs(min_x - inner_min_x) <= tolerance:
        contacts += 1
    if abs(min_y - inner_min_y) <= tolerance:
        contacts += 1
    if abs(max_x - inner_max_x) <= tolerance:
        contacts += 1
    if abs(max_y - inner_max_y) <= tolerance:
        contacts += 1

    for other in placed_polygons:
        if not bounds_are_close(
            polygon.bounds,
            other.bounds,
            clearance + tolerance,
        ):
            continue
        if polygon.distance(other) <= clearance + tolerance:
            contacts += 1

    return contacts


def placement_score(
    polygon: Polygon,
    current_max_x: float,
    current_max_y: float,
    current_material_area: float,
    placed_polygons: list[Polygon],
    sheet_inner: Polygon,
    clearance: float,
    margin: float,
) -> tuple:
    min_x, min_y, max_x, max_y = polygon.bounds
    new_max_x = max(current_max_x, max_x)
    new_max_y = max(current_max_y, max_y)
    envelope_width = max(1.0, new_max_x - margin)
    envelope_height = max(1.0, new_max_y - margin)
    envelope_area = envelope_width * envelope_height
    material_area = current_material_area + polygon.area
    waste = max(0.0, envelope_area - material_area)
    waste_ratio = waste / envelope_area
    contacts = contact_count(
        polygon,
        placed_polygons,
        sheet_inner,
        clearance,
    )

    center_x = (min_x + max_x) / 2.0
    center_y = (min_y + max_y) / 2.0
    sheet_min_x, sheet_min_y, sheet_max_x, sheet_max_y = sheet_inner.bounds
    distance_to_edges = min(
        abs(min_x - sheet_min_x),
        abs(min_y - sheet_min_y),
        abs(sheet_max_x - max_x),
        abs(sheet_max_y - max_y),
    )

    if is_height_priority():
        # Priorité au remplissage des 1500 mm :
        # on garde la largeur consommée la plus faible possible,
        # afin de remplir une colonne avant d'ouvrir la suivante.
        return (
            round(waste_ratio, 7),
            round(waste, 1),
            -contacts * 3,
            round(distance_to_edges, 2),
            round(envelope_area, 1),
            round(new_max_x, 3),
            round(new_max_y, 3),
            round(center_x, 3),
            round(center_y, 3),
        )

    if is_width_priority():
        return (
            round(waste_ratio, 7),
            round(waste, 1),
            -contacts * 3,
            round(distance_to_edges, 2),
            round(envelope_area, 1),
            round(new_max_y, 3),
            round(new_max_x, 3),
            round(center_y, 3),
            round(center_x, 3),
        )

    balance = abs((new_max_x - margin) - (new_max_y - margin))
    return (
        round(waste_ratio, 7),
        round(waste, 1),
        -contacts * 3,
        round(distance_to_edges, 2),
        round(envelope_area, 1),
        round(balance, 3),
        round(new_max_y, 3),
        round(new_max_x, 3),
    )


def expand_piece_copies(
    pieces: list[DxfPiece],
) -> list[tuple[DxfPiece, int]]:
    expanded: list[tuple[DxfPiece, int]] = []

    for piece in pieces:
        for copy_index in range(1, piece.quantity + 1):
            expanded.append((piece, copy_index))

    return expanded


def order_piece_copies(
    expanded: list[tuple[DxfPiece, int]],
    attempt_index: int,
) -> list[tuple[DxfPiece, int]]:
    ordered = list(expanded)
    strategy = attempt_index % 16

    def bbox_area(item):
        polygon = item[0].polygon
        min_x, min_y, max_x, max_y = polygon.bounds
        return (max_x - min_x) * (max_y - min_y)



def compactness_score(
    sheets: list[list[Placement]],
    margin: float,
) -> tuple:
    total_waste = 0.0
    total_envelope = 0.0
    total_height = 0.0

    for sheet in sheets:
        if not sheet:
            continue

        max_x = max(item.polygon.bounds[2] for item in sheet)
        max_y = max(item.polygon.bounds[3] for item in sheet)
        material_area = sum(item.polygon.area for item in sheet)
        envelope = max(1.0, max_x - margin) * max(1.0, max_y - margin)
        total_envelope += envelope
        total_waste += max(0.0, envelope - material_area)
        total_height += max_y - margin

    total_width = 0.0
    for sheet in sheets:
        if not sheet:
            continue
        max_x = max(item.polygon.bounds[2] for item in sheet)
        total_width += max_x - margin

    direction_metric = (
        total_width
        if is_height_priority()
        else total_height
        if is_width_priority()
        else total_width + total_height
    )

    return (
        len(sheets),
        round(total_waste, 1),
        round(total_envelope, 1),
        round(direction_metric, 1),
    )


def fallback_shelf_nest(
    pieces: list[DxfPiece],
    sheet_width: float,
    sheet_height: float,
    margin: float,
    clearance: float,
    rotations: list[int],
    rotation_cache: dict[tuple[str, int], Polygon],
) -> tuple[list[Placement], list[list[Placement]], list[str]]:
    inner_right = sheet_width - margin
    inner_top = sheet_height - margin
    expanded = expand_piece_copies(pieces)
    expanded.sort(
        key=lambda item: (
            max(
                item[0].polygon.bounds[2] - item[0].polygon.bounds[0],
                item[0].polygon.bounds[3] - item[0].polygon.bounds[1],
            ),
            item[0].polygon.area,
        ),
        reverse=True,
    )

    simple_angles = [angle for angle in (0, 90, 180, 270) if angle in rotations]
    if not simple_angles:
        simple_angles = [rotations[0]]

    sheets_data: list[dict] = []
    unplaced: list[str] = []

    for piece, copy_index in expanded:
        versions = []
        for angle in simple_angles:
            cache_key = (piece.reference_key, angle)
            if cache_key not in rotation_cache:
                rotation_cache[cache_key] = rotated_at_origin(piece.polygon, angle)
            polygon = rotation_cache[cache_key]
            width = polygon.bounds[2] - polygon.bounds[0]
            height = polygon.bounds[3] - polygon.bounds[1]
            versions.append((angle, polygon, width, height))

        best = None
        for sheet_index, data in enumerate(sheets_data):
            for shelf_index, shelf in enumerate(data['shelves']):
                for angle, polygon, width, height in versions:
                    if (
                        height <= shelf['height'] + 1e-6
                        and shelf['x'] + width <= inner_right + 1e-6
                    ):
                        remaining = inner_right - (shelf['x'] + width)
                        score = (0, remaining, shelf['y'], sheet_index)
                        if best is None or score < best[0]:
                            best = (score, 'existing', sheet_index, shelf_index, angle, polygon, width, height)

            new_y = margin if not data['shelves'] else max(
                shelf['y'] + shelf['height'] + clearance
                for shelf in data['shelves']
            )
            for angle, polygon, width, height in versions:
                if margin + width <= inner_right + 1e-6 and new_y + height <= inner_top + 1e-6:
                    score = (1, new_y + height, width, sheet_index)
                    if best is None or score < best[0]:
                        best = (score, 'new', sheet_index, None, angle, polygon, width, height)

        if best is None:
            valid = []
            for angle, polygon, width, height in versions:
                if margin + width <= inner_right + 1e-6 and margin + height <= inner_top + 1e-6:
                    valid.append((height * width, height, width, angle, polygon))
            if not valid:
                unplaced.append(f'{piece.reference_display} - copie {copy_index}')
                continue
            _, height, width, angle, polygon = min(valid)
            sheet_index = len(sheets_data)
            sheets_data.append({
                'placements': [],
                'shelves': [{'y': margin, 'height': height, 'x': margin}],
            })
            best = ((2, height, width, sheet_index), 'existing', sheet_index, 0, angle, polygon, width, height)

        _, option, sheet_index, shelf_index, angle, polygon, width, height = best
        data = sheets_data[sheet_index]
        if option == 'new':
            new_y = margin if not data['shelves'] else max(
                shelf['y'] + shelf['height'] + clearance
                for shelf in data['shelves']
            )
            data['shelves'].append({'y': new_y, 'height': height, 'x': margin})
            shelf_index = len(data['shelves']) - 1

        shelf = data['shelves'][shelf_index]
        placed_polygon = affinity.translate(polygon, xoff=shelf['x'], yoff=shelf['y'])
        data['placements'].append(
            Placement(
                reference=piece.reference_display,
                source_name=piece.source_name,
                sheet_index=sheet_index,
                rotation=angle,
                polygon=placed_polygon,
                original_polygon=piece.polygon,
                copy_index=copy_index,
                thickness=piece.thickness,
                material=piece.material,
            )
        )
        shelf['x'] += width + clearance
        shelf['height'] = max(shelf['height'], height)

    sheets = [data['placements'] for data in sheets_data]
    placements = [item for sheet in sheets for item in sheet]
    return placements, sheets, unplaced


def nest_one_order(
    ordered_pieces: list[tuple[DxfPiece, int]],
    sheet_width: float,
    sheet_height: float,
    margin: float,
    clearance: float,
    rotations: list[int],
    allow_hole_nesting: bool,
    quality_level: int,
    max_candidates: int,
    rotation_cache: dict[tuple[str, int], Polygon],
    deadline: float,
    item_progress_callback=None,
    sheet_limit: int | None = None,
) -> tuple[list[Placement], list[list[Placement]], list[str], bool]:
    sheet_inner = box(
        margin,
        margin,
        sheet_width - margin,
        sheet_height - margin,
    )
    inner_width = sheet_width - 2 * margin
    inner_height = sheet_height - 2 * margin
    sheets: list[list[Placement]] = []
    unplaced: list[str] = []
    total = len(ordered_pieces)

    for item_index, (piece, copy_index) in enumerate(ordered_pieces, start=1):
        if time.perf_counter() >= deadline:
            return [], [], [], True

        versions = []
        for angle in piece_rotation_angles(piece, rotations):
            cache_key = (piece.reference_key, angle)
            if cache_key not in rotation_cache:
                rotation_cache[cache_key] = rotated_at_origin(piece.polygon, angle)
            versions.append((angle, rotation_cache[cache_key]))

        best_global = None

        for sheet_index, sheet_placements in enumerate(sheets):
            if time.perf_counter() >= deadline:
                return [], [], [], True

            fixed = [item.polygon for item in sheet_placements]
            current_max_x = max((polygon.bounds[2] for polygon in fixed), default=margin)
            current_max_y = max((polygon.bounds[3] for polygon in fixed), default=margin)
            current_area = sum(polygon.area for polygon in fixed)

            for angle, rotated in versions:
                width = rotated.bounds[2] - rotated.bounds[0]
                height = rotated.bounds[3] - rotated.bounds[1]
                if width > inner_width + 1e-6 or height > inner_height + 1e-6:
                    continue

                positions = candidate_positions(
                    rotated,
                    fixed,
                    margin,
                    clearance,
                    sheet_width,
                    sheet_height,
                    allow_hole_nesting,
                    quality_level,
                    max_candidates,
                )

                for x, y in positions:
                    if time.perf_counter() >= deadline:
                        return [], [], [], True
                    candidate = affinity.translate(rotated, xoff=x, yoff=y)
                    if not fits_on_sheet(candidate, fixed, sheet_inner, clearance):
                        continue
                    if quality_level >= 2:
                        candidate = directional_compact(
                            candidate,
                            fixed,
                            sheet_inner,
                            clearance,
                            minimum_step=1.0 if quality_level == 2 else 0.5,
                        )

                    local = placement_score(
                        candidate,
                        current_max_x,
                        current_max_y,
                        current_area,
                        fixed,
                        sheet_inner,
                        clearance,
                        margin,
                    )
                    global_score = (*local, sheet_index)
                    if best_global is None or global_score < best_global[0]:
                        best_global = (global_score, sheet_index, angle, candidate)

        if best_global is not None:
            _, sheet_index, angle, candidate = best_global
            sheets[sheet_index].append(
                Placement(
                    reference=piece.reference_display,
                    source_name=piece.source_name,
                    sheet_index=sheet_index,
                    rotation=angle,
                    polygon=candidate,
                    original_polygon=piece.polygon,
                    copy_index=copy_index,
                    thickness=piece.thickness,
                    material=piece.material,
                )
            )
        else:
            if sheet_limit is not None and len(sheets) >= sheet_limit:
                unplaced.append(f'{piece.reference_display} - copie {copy_index}')
            else:
                best_new = None
                for angle, rotated in versions:
                    width = rotated.bounds[2] - rotated.bounds[0]
                    height = rotated.bounds[3] - rotated.bounds[1]
                    if width > inner_width + 1e-6 or height > inner_height + 1e-6:
                        continue
                    candidate = affinity.translate(rotated, xoff=margin, yoff=margin)
                    if sheet_inner.covers(candidate):
                        envelope = width * height
                        waste_ratio = max(0.0, envelope - candidate.area) / max(1.0, envelope)
                        score = (waste_ratio, envelope, height, width)
                        if best_new is None or score < best_new[0]:
                            best_new = (score, angle, candidate)

                if best_new is None:
                    unplaced.append(f'{piece.reference_display} - copie {copy_index}')
                else:
                    _, angle, candidate = best_new
                    sheet_index = len(sheets)
                    sheets.append([
                        Placement(
                            reference=piece.reference_display,
                            source_name=piece.source_name,
                            sheet_index=sheet_index,
                            rotation=angle,
                            polygon=candidate,
                            original_polygon=piece.polygon,
                            copy_index=copy_index,
                            thickness=piece.thickness,
                            material=piece.material,
                        )
                    ])

        if item_progress_callback is not None and total:
            item_progress_callback(item_index / total)

    placements = [item for sheet in sheets for item in sheet]
    return placements, sheets, unplaced, False


def placement_to_piece(item: Placement) -> DxfPiece:
    return DxfPiece(
        reference_display=item.reference,
        reference_key=normalize_reference(item.reference),
        source_name=item.source_name,
        polygon=item.original_polygon,
        quantity=1,
        thickness=item.thickness,
        material=item.material,
    )


def try_insert_item(
    item: Placement,
    target_sheets: list[list[Placement]],
    sheet_width: float,
    sheet_height: float,
    margin: float,
    clearance: float,
    rotations: list[int],
    allow_hole_nesting: bool,
    rotation_cache: dict[tuple[str, int], Polygon],
    deadline: float,
    quality_level: int = 3,
    max_candidates: int = 180,
):
    sheet_inner = box(margin, margin, sheet_width - margin, sheet_height - margin)
    reference_key = normalize_reference(item.reference)
    best = None

    for target_index, target in enumerate(target_sheets):
        if time.perf_counter() >= deadline:
            return None
        fixed = [placed.polygon for placed in target]
        current_max_x = max((p.bounds[2] for p in fixed), default=margin)
        current_max_y = max((p.bounds[3] for p in fixed), default=margin)
        current_area = sum(p.area for p in fixed)

        item_piece = piece_like_from_placement(item)
        for angle in piece_rotation_angles(item_piece, rotations):
            cache_key = (reference_key, angle)
            if cache_key not in rotation_cache:
                rotation_cache[cache_key] = rotated_at_origin(item.original_polygon, angle)
            rotated = rotation_cache[cache_key]
            positions = candidate_positions(
                rotated,
                fixed,
                margin,
                clearance,
                sheet_width,
                sheet_height,
                allow_hole_nesting,
                quality_level,
                max_candidates,
            )
            for x, y in positions:
                candidate = affinity.translate(rotated, xoff=x, yoff=y)
                if not fits_on_sheet(candidate, fixed, sheet_inner, clearance):
                    continue
                candidate = directional_compact(
                    candidate,
                    fixed,
                    sheet_inner,
                    clearance,
                    minimum_step=0.5,
                )
                score = placement_score(
                    candidate,
                    current_max_x,
                    current_max_y,
                    current_area,
                    fixed,
                    sheet_inner,
                    clearance,
                    margin,
                )
                global_score = (*score, target_index)
                if best is None or global_score < best[0]:
                    best = (global_score, target_index, angle, candidate)

    return best


def consolidate_sparse_sheets(
    sheets: list[list[Placement]],
    sheet_width: float,
    sheet_height: float,
    margin: float,
    clearance: float,
    rotations: list[int],
    allow_hole_nesting: bool,
    rotation_cache: dict[tuple[str, int], Polygon],
    deadline: float,
) -> list[list[Placement]]:
    """Essaie de supprimer toutes les tôles les moins remplies, pas seulement la dernière."""
    usable_area = max(1.0, (sheet_width - 2 * margin) * (sheet_height - 2 * margin))
    changed = True

    while changed and len(sheets) > 1 and time.perf_counter() < deadline:
        changed = False
        source_order = sorted(
            range(len(sheets)),
            key=lambda index: sum(item.polygon.area for item in sheets[index]) / usable_area,
        )

        for source_index in source_order:
            if time.perf_counter() >= deadline:
                break
            source_items = sorted(
                sheets[source_index],
                key=lambda item: item.original_polygon.area,
                reverse=True,
            )
            trial = [list(sheet) for index, sheet in enumerate(sheets) if index != source_index]
            success = True

            for item in source_items:
                best = try_insert_item(
                    item,
                    trial,
                    sheet_width,
                    sheet_height,
                    margin,
                    clearance,
                    rotations,
                    allow_hole_nesting,
                    rotation_cache,
                    deadline,
                )
                if best is None:
                    success = False
                    break
                _, target_index, angle, candidate = best
                trial[target_index].append(
                    Placement(
                        reference=item.reference,
                        source_name=item.source_name,
                        sheet_index=target_index,
                        rotation=angle,
                        polygon=candidate,
                        original_polygon=item.original_polygon,
                        copy_index=item.copy_index,
                        thickness=item.thickness,
                        material=item.material,
                    )
                )

            if success:
                sheets = trial
                changed = True
                break

    for sheet_index, sheet in enumerate(sheets):
        for placement in sheet:
            placement.sheet_index = sheet_index
    return sheets


def repack_sparse_pairs(
    sheets: list[list[Placement]],
    sheet_width: float,
    sheet_height: float,
    margin: float,
    clearance: float,
    rotations: list[int],
    allow_hole_nesting: bool,
    rotation_cache: dict[tuple[str, int], Polygon],
    deadline: float,
) -> list[list[Placement]]:
    """Tente de fusionner les deux tôles les moins chargées en une seule."""
    while len(sheets) > 1 and time.perf_counter() < deadline:
        utilization = sorted(
            range(len(sheets)),
            key=lambda index: sum(item.polygon.area for item in sheets[index]),
        )
        pair = utilization[:2]
        collected = [item for index in pair for item in sheets[index]]
        expanded = [(placement_to_piece(item), item.copy_index) for item in collected]
        best_single = None

        for attempt in range(4):
            if time.perf_counter() >= deadline:
                break
            ordered = order_piece_copies(expanded, attempt)
            placements, candidate_sheets, unplaced, timed_out = nest_one_order(
                ordered,
                sheet_width,
                sheet_height,
                margin,
                clearance,
                rotations,
                allow_hole_nesting,
                3,
                220,
                rotation_cache,
                deadline,
                sheet_limit=1,
            )
            if timed_out:
                break
            if not unplaced and len(candidate_sheets) == 1:
                score = compactness_score(candidate_sheets, margin)
                if best_single is None or score < best_single[0]:
                    best_single = (score, candidate_sheets[0])

        if best_single is None:
            break

        remaining = [sheet for index, sheet in enumerate(sheets) if index not in pair]
        remaining.append(best_single[1])
        sheets = remaining

    for sheet_index, sheet in enumerate(sheets):
        for placement in sheet:
            placement.sheet_index = sheet_index
    return sheets


def group_complexity_weight(pieces: list[DxfPiece]) -> float:
    """
    Estime le temps nécessaire à un groupe indépendamment de son épaisseur.

    Le poids tient compte :
    - du nombre de copies ;
    - du nombre de sommets ;
    - des trous ;
    - de la concavité des contours ;
    - de la diversité des repères.
    """
    weight = 0.0

    for piece in pieces:
        polygon = piece.polygon
        copies = max(1, piece.quantity)
        min_x, min_y, max_x, max_y = polygon.bounds
        bounding_area = max(1.0, (max_x - min_x) * (max_y - min_y))
        concavity = min(6.0, bounding_area / max(1.0, polygon.area))
        vertices = len(polygon.exterior.coords)
        holes = len(polygon.interiors)

        geometry_factor = (
            1.0
            + min(2.0, vertices / 35.0)
            + min(2.0, holes * 0.35)
            + min(2.0, (concavity - 1.0) * 0.45)
        )

        weight += copies * geometry_factor

    # Petit bonus de temps lorsque le groupe comporte beaucoup de formes distinctes.
    weight += len(pieces) * 1.5
    return max(1.0, weight)


def repack_sparse_clusters(
    sheets: list[list[Placement]],
    sheet_width: float,
    sheet_height: float,
    margin: float,
    clearance: float,
    rotations: list[int],
    allow_hole_nesting: bool,
    rotation_cache: dict[tuple[str, int], Polygon],
    deadline: float,
    cluster_size: int = 3,
    target_sheet_count: int = 2,
) -> list[list[Placement]]:
    """
    Sélectionne les tôles les moins remplies et tente de replacer toutes
    leurs pièces sur un nombre inférieur de tôles.

    Cette étape est générale : elle s'applique à toutes les épaisseurs.
    """
    if cluster_size <= target_sheet_count:
        return sheets

    while (
        len(sheets) >= cluster_size
        and time.perf_counter() < deadline
    ):
        least_filled = sorted(
            range(len(sheets)),
            key=lambda index: sum(
                item.polygon.area for item in sheets[index]
            ),
        )[:cluster_size]

        collected = [
            item
            for index in least_filled
            for item in sheets[index]
        ]

        expanded = [
            (placement_to_piece(item), item.copy_index)
            for item in collected
        ]

        best_candidate = None

        for attempt in range(8):
            if time.perf_counter() >= deadline:
                break

            ordered = order_piece_copies(expanded, attempt)

            placements, candidate_sheets, unplaced, timed_out = nest_one_order(
                ordered,
                sheet_width,
                sheet_height,
                margin,
                clearance,
                rotations,
                allow_hole_nesting,
                3,
                300,
                rotation_cache,
                deadline,
                sheet_limit=target_sheet_count,
            )

            if timed_out:
                break

            if (
                not unplaced
                and len(candidate_sheets) <= target_sheet_count
            ):
                score = compactness_score(candidate_sheets, margin)

                if (
                    best_candidate is None
                    or score < best_candidate[0]
                ):
                    best_candidate = (
                        score,
                        candidate_sheets,
                    )

        if best_candidate is None:
            break

        remaining = [
            sheet
            for index, sheet in enumerate(sheets)
            if index not in least_filled
        ]
        remaining.extend(best_candidate[1])
        sheets = remaining

    for sheet_index, sheet in enumerate(sheets):
        for placement in sheet:
            placement.sheet_index = sheet_index

    return sheets



def nest_pieces(
    pieces: list[DxfPiece],
    sheet_width: float,
    sheet_height: float,
    margin: float,
    clearance: float,
    rotations: list[int],
    quality_mode: str = 'Équilibré',
    allow_hole_nesting: bool = True,
    time_budget_seconds: float = 90.0,
    progress_callback=None,
) -> tuple[list[Placement], int, list[str]]:
    if sheet_width <= 0 or sheet_height <= 0:
        raise ValueError('Les dimensions de la tôle doivent être positives.')
    if margin < 0 or clearance < 0:
        raise ValueError("La marge et l'espacement ne peuvent pas être négatifs.")
    if sheet_width - 2 * margin <= 0 or sheet_height - 2 * margin <= 0:
        raise ValueError('La marge est trop grande pour le format de tôle.')
    if not rotations:
        rotations = [0]

    expanded = expand_piece_copies(pieces)
    total_copies = len(expanded)
    rotation_cache: dict[tuple[str, int], Polygon] = {}
    effective_rotations = sorted(set(rotations))

    quality_settings = {
        'Rapide': (2, 1, 70),
        'Équilibré': (7, 2, 150),
        'Approfondi': (15, 3, 270),
        'Maximum': (30, 3, 420),
    }
    attempts, quality_level, max_candidates = quality_settings.get(
        quality_mode,
        (6, 2, 130),
    )

    if quality_mode == "Maximum":
        if total_copies > 220:
            attempts = min(attempts, 12)
            max_candidates = min(max_candidates, 260)
        elif total_copies > 140:
            attempts = min(attempts, 18)
            max_candidates = min(max_candidates, 340)
    else:
        if total_copies > 180:
            attempts = min(attempts, 5)
            max_candidates = min(max_candidates, 140)
        elif total_copies > 100:
            attempts = min(attempts, 9)
            max_candidates = min(max_candidates, 200)

    # On conserve les rotations à 180° même pour les gros projets : elles
    # permettent aux formes coudées de s'emboîter tête-bêche.
    if total_copies > 160 and len(effective_rotations) > 4:
        effective_rotations = [angle for angle in (0, 90, 180, 270) if angle in effective_rotations]
        if not effective_rotations:
            effective_rotations = sorted(set(rotations))[:4]

    fallback_placements, fallback_sheets, fallback_unplaced = fallback_shelf_nest(
        pieces,
        sheet_width,
        sheet_height,
        margin,
        clearance,
        effective_rotations,
        rotation_cache,
    )
    best_result = (
        (len(fallback_unplaced), *compactness_score(fallback_sheets, margin)),
        fallback_placements,
        fallback_sheets,
        fallback_unplaced,
    )

    if progress_callback is not None:
        progress_callback(0.05)

    deadline = time.perf_counter() + max(10.0, float(time_budget_seconds))

    for attempt_index in range(attempts):
        if time.perf_counter() >= deadline:
            break
        ordered = order_piece_copies(expanded, attempt_index)

        def update_item_progress(item_fraction):
            if progress_callback is not None:
                fraction = 0.05 + 0.72 * (attempt_index + item_fraction) / max(1, attempts)
                progress_callback(min(0.77, fraction))

        placements, sheets, unplaced, timed_out = nest_one_order(
            ordered,
            sheet_width,
            sheet_height,
            margin,
            clearance,
            effective_rotations,
            allow_hole_nesting,
            quality_level,
            max_candidates,
            rotation_cache,
            deadline,
            item_progress_callback=update_item_progress,
        )
        if timed_out:
            break

        score = (len(unplaced), *compactness_score(sheets, margin))
        if score < best_result[0]:
            best_result = (score, placements, sheets, unplaced)

    _, placements, sheets, unplaced = best_result

    if not unplaced and len(sheets) > 1 and time.perf_counter() < deadline:
        if progress_callback is not None:
            progress_callback(0.80)
        sheets = consolidate_sparse_sheets(
            sheets,
            sheet_width,
            sheet_height,
            margin,
            clearance,
            effective_rotations,
            allow_hole_nesting,
            rotation_cache,
            deadline,
        )

    if not unplaced and len(sheets) > 1 and time.perf_counter() < deadline:
        if progress_callback is not None:
            progress_callback(0.86)
        sheets = repack_sparse_pairs(
            sheets,
            sheet_width,
            sheet_height,
            margin,
            clearance,
            effective_rotations,
            allow_hole_nesting,
            rotation_cache,
            deadline,
        )

    if not unplaced and len(sheets) >= 3 and time.perf_counter() < deadline:
        if progress_callback is not None:
            progress_callback(0.92)
        sheets = repack_sparse_clusters(
            sheets,
            sheet_width,
            sheet_height,
            margin,
            clearance,
            effective_rotations,
            allow_hole_nesting,
            rotation_cache,
            deadline,
            cluster_size=3,
            target_sheet_count=2,
        )

    if (
        quality_mode == "Maximum"
        and not unplaced
        and len(sheets) >= 4
        and time.perf_counter() < deadline
    ):
        if progress_callback is not None:
            progress_callback(0.96)
        sheets = repack_sparse_clusters(
            sheets,
            sheet_width,
            sheet_height,
            margin,
            clearance,
            effective_rotations,
            allow_hole_nesting,
            rotation_cache,
            deadline,
            cluster_size=4,
            target_sheet_count=3,
        )

    placements = [placement for sheet in sheets for placement in sheet]
    for sheet_index, sheet in enumerate(sheets):
        for placement in sheet:
            placement.sheet_index = sheet_index

    if progress_callback is not None:
        progress_callback(1.0)

    return placements, len(sheets), unplaced


# ============================================================
# V13 — Stock de tôles à dimensions variables
# ============================================================


# ============================================================
# Correctifs géométriques globaux v18
# ============================================================

def _geometry_from_any_item(item):
    """
    Récupère le polygone depuis :
    - un tuple (DxfPiece, copie)
    - un DxfPiece
    - un Placement
    """
    if isinstance(item, tuple) and item:
        first = item[0]
        if hasattr(first, "polygon"):
            return first.polygon
        if hasattr(first, "original_polygon"):
            return first.original_polygon

    if hasattr(item, "polygon"):
        return item.polygon

    if hasattr(item, "original_polygon"):
        return item.original_polygon

    raise ValueError("Objet géométrique non reconnu.")


def max_dimension(item) -> float:
    polygon = _geometry_from_any_item(item)
    min_x, min_y, max_x, max_y = polygon.bounds
    return max(max_x - min_x, max_y - min_y)


def min_dimension(item) -> float:
    polygon = _geometry_from_any_item(item)
    min_x, min_y, max_x, max_y = polygon.bounds
    return min(max_x - min_x, max_y - min_y)


def concavity(item) -> float:
    polygon = _geometry_from_any_item(item)
    min_x, min_y, max_x, max_y = polygon.bounds
    bounding_area = max(1.0, (max_x - min_x) * (max_y - min_y))
    polygon_area = max(1.0, polygon.area)
    return bounding_area / polygon_area



@dataclass
class StockSheet:
    instance_id: str
    row_id: int
    stock_id: str
    material: str
    thickness: str
    width: float
    height: float
    stock_type: str
    priority: int


@dataclass
class SheetMeta:
    sheet_index: int
    sheet_id: str
    material: str
    thickness: str
    width: float
    height: float
    source: str
    stock_type: str
    stock_row_id: int | None
    priority: int


STOCK_COLUMNS = [
    "Utiliser",
    "ID stock",
    "Matière",
    "Épaisseur",
    "Largeur (mm)",
    "Hauteur (mm)",
    "Quantité",
    "Type",
    "Priorité",
]


def default_stock_dataframe(
    default_width: float = 3000.0,
    default_height: float = 1500.0,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Utiliser": True,
                "ID stock": "STOCK-001",
                "Matière": "S235JR",
                "Épaisseur": "5",
                "Largeur (mm)": float(default_width),
                "Hauteur (mm)": float(default_height),
                "Quantité": 1,
                "Type": "Tôle complète",
                "Priorité": 10,
            }
        ],
        columns=STOCK_COLUMNS,
    )


def ensure_stock_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    result = dataframe.copy()

    defaults = {
        "Utiliser": True,
        "ID stock": "",
        "Matière": "",
        "Épaisseur": "",
        "Largeur (mm)": 3000.0,
        "Hauteur (mm)": 1500.0,
        "Quantité": 1,
        "Type": "Tôle complète",
        "Priorité": 10,
    }

    for column in STOCK_COLUMNS:
        if column not in result.columns:
            result[column] = defaults[column]

    return result[STOCK_COLUMNS]


def read_stock_csv(uploaded_file) -> pd.DataFrame:
    raw = uploaded_file.getvalue()
    decoded = None

    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            decoded = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if decoded is None:
        raise ValueError("Impossible de lire le CSV de stock.")

    try:
        dataframe = pd.read_csv(
            io.StringIO(decoded),
            sep=None,
            engine="python",
        )
    except Exception as exc:
        raise ValueError(
            "Le CSV de stock est illisible. "
            "Utilise un séparateur point-virgule ou virgule."
        ) from exc

    return ensure_stock_columns(dataframe)


def material_key(value: object) -> str:
    text = remove_accents(cell_to_text(value)).upper()
    return re.sub(r"[^A-Z0-9*]+", "", text)


def thickness_number(value: object) -> float | None:
    text = cell_to_text(value).replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)

    if not match:
        return None

    try:
        return float(match.group(0))
    except ValueError:
        return None


def thickness_key(value: object) -> str:
    number = thickness_number(value)

    if number is not None:
        return f"{number:.4f}"

    return normalize_reference(value)


def stock_matches_group(
    stock_sheet: StockSheet,
    material: str,
    thickness: str,
) -> bool:
    stock_material = material_key(stock_sheet.material)
    group_material = material_key(material)

    wildcard_materials = {"*", "TOUTES", "TOUS", "ALL"}
    unknown_materials = {"", "NONRENSEIGNEE", "INCONNUE"}

    if stock_material in wildcard_materials:
        material_match = True
    elif group_material in unknown_materials:
        material_match = stock_material in unknown_materials
    else:
        material_match = stock_material == group_material

    stock_thickness_number = thickness_number(stock_sheet.thickness)
    group_thickness_number = thickness_number(thickness)

    if (
        stock_thickness_number is not None
        and group_thickness_number is not None
    ):
        thickness_match = (
            abs(stock_thickness_number - group_thickness_number) <= 0.01
        )
    else:
        thickness_match = (
            thickness_key(stock_sheet.thickness)
            == thickness_key(thickness)
        )

    return material_match and thickness_match


def prepare_stock_instances(
    stock_dataframe: pd.DataFrame,
) -> list[StockSheet]:
    dataframe = ensure_stock_columns(stock_dataframe)
    instances: list[StockSheet] = []

    for row_id, row in dataframe.reset_index(drop=True).iterrows():
        use_value = row.get("Utiliser", True)

        if isinstance(use_value, str):
            enabled = use_value.strip().lower() not in {
                "non",
                "false",
                "0",
                "désactivé",
                "desactive",
            }
        else:
            enabled = bool(use_value)

        if not enabled:
            continue

        try:
            width = float(str(row["Largeur (mm)"]).replace(",", "."))
            height = float(str(row["Hauteur (mm)"]).replace(",", "."))
            quantity = int(float(str(row["Quantité"]).replace(",", ".")))
            priority = int(float(str(row["Priorité"]).replace(",", ".")))
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Valeur incorrecte dans le stock à la ligne {row_id + 1}."
            ) from exc

        if width <= 0 or height <= 0 or quantity <= 0:
            continue

        stock_id = cell_to_text(row["ID stock"]) or f"STOCK-{row_id + 1:03d}"
        material = cell_to_text(row["Matière"])
        thickness = cell_to_text(row["Épaisseur"])
        stock_type = cell_to_text(row["Type"]) or "Tôle complète"

        if not thickness:
            raise ValueError(
                f"Épaisseur manquante dans le stock à la ligne {row_id + 1}."
            )

        for copy_number in range(1, quantity + 1):
            instances.append(
                StockSheet(
                    instance_id=f"{stock_id}-{copy_number:03d}",
                    row_id=int(row_id),
                    stock_id=stock_id,
                    material=material,
                    thickness=thickness,
                    width=width,
                    height=height,
                    stock_type=stock_type,
                    priority=priority,
                )
            )

    return instances


def copy_key(piece: DxfPiece, copy_index: int) -> tuple[str, int]:
    return piece.reference_key, int(copy_index)


def placement_key(placement: Placement) -> tuple[str, int]:
    return normalize_reference(placement.reference), int(placement.copy_index)


def remove_placed_copies(
    remaining: list[tuple[DxfPiece, int]],
    placements: list[Placement],
) -> list[tuple[DxfPiece, int]]:
    used_keys = {placement_key(item) for item in placements}

    return [
        (piece, copy_index)
        for piece, copy_index in remaining
        if copy_key(piece, copy_index) not in used_keys
    ]


def stock_orderings(
    stock_sheets: list[StockSheet],
    objective: str,
) -> list[list[StockSheet]]:
    def type_rank(item: StockSheet) -> int:
        return 0 if "CHUTE" in material_key(item.stock_type) else 1

    orderings = [
        sorted(
            stock_sheets,
            key=lambda item: (
                item.priority,
                type_rank(item),
                item.width * item.height,
            ),
        ),
        sorted(
            stock_sheets,
            key=lambda item: (
                item.priority,
                type_rank(item),
                -(item.width * item.height),
            ),
        ),
        sorted(
            stock_sheets,
            key=lambda item: (
                type_rank(item),
                item.width * item.height,
                item.priority,
            ),
        ),
    ]

    if objective == "Minimiser le nombre de plaques":
        orderings.insert(
            0,
            sorted(
                stock_sheets,
                key=lambda item: (
                    item.priority,
                    -(item.width * item.height),
                ),
            ),
        )

    unique = []
    signatures = set()

    for ordering in orderings:
        signature = tuple(item.instance_id for item in ordering)

        if signature not in signatures:
            signatures.add(signature)
            unique.append(ordering)

    return unique


def quality_parameters(quality_mode: str) -> tuple[int, int, int]:
    settings = {
        "Rapide": (1, 1, 70),
        "Équilibré": (3, 2, 150),
        "Approfondi": (7, 3, 280),
        "Maximum": (12, 3, 420),
        "Pro dense": (18, 3, 650),
    }
    return settings.get(quality_mode, settings["Équilibré"])



def vertical_column_fill_one_sheet(
    remaining: list[tuple[DxfPiece, int]],
    width: float,
    height: float,
    margin: float,
    clearance: float,
    rotations: list[int],
    rotation_cache: dict[tuple[str, int], Polygon],
    deadline: float,
) -> list[Placement]:
    """
    Remplit la plaque par colonnes : d'abord la hauteur, puis la largeur.
    """
    if not remaining:
        return []

    inner_right = width - margin
    inner_top = height - margin
    current_x = margin
    current_y = margin
    column_width = 0.0
    placements: list[Placement] = []

    ordered = sorted(
        remaining,
        key=lambda item: (
            max_dimension(item),
            item[0].polygon.area,
        ),
        reverse=True,
    )

    for piece, copy_index in ordered:
        if time.perf_counter() >= deadline:
            break

        versions = []

        for angle in piece_rotation_angles(piece, rotations):
            cache_key = (piece.reference_key, angle)

            if cache_key not in rotation_cache:
                rotation_cache[cache_key] = rotated_at_origin(
                    piece.polygon,
                    angle,
                )

            polygon = rotation_cache[cache_key]
            piece_width = polygon.bounds[2] - polygon.bounds[0]
            piece_height = polygon.bounds[3] - polygon.bounds[1]

            if (
                piece_width <= width - 2 * margin + 1e-6
                and piece_height <= height - 2 * margin + 1e-6
            ):
                versions.append(
                    (
                        angle,
                        polygon,
                        piece_width,
                        piece_height,
                    )
                )

        if not versions:
            continue

        versions.sort(
            key=lambda item: (
                item[2],
                -item[3],
            )
        )

        placed = False

        for angle, polygon, piece_width, piece_height in versions:
            if (
                current_x + max(column_width, piece_width)
                <= inner_right + 1e-6
                and current_y + piece_height <= inner_top + 1e-6
            ):
                candidate = affinity.translate(
                    polygon,
                    xoff=current_x,
                    yoff=current_y,
                )
                placements.append(
                    Placement(
                        reference=piece.reference_display,
                        source_name=piece.source_name,
                        sheet_index=0,
                        rotation=angle,
                        polygon=candidate,
                        original_polygon=piece.polygon,
                        copy_index=copy_index,
                        thickness=piece.thickness,
                        material=piece.material,
                    )
                )
                current_y += piece_height + clearance
                column_width = max(column_width, piece_width)
                placed = True
                break

        if placed:
            continue

        new_x = current_x + column_width + clearance
        best_new_column = None

        for angle, polygon, piece_width, piece_height in versions:
            if (
                new_x + piece_width <= inner_right + 1e-6
                and margin + piece_height <= inner_top + 1e-6
            ):
                score = (
                    piece_width,
                    -piece_height,
                )
                if best_new_column is None or score < best_new_column[0]:
                    best_new_column = (
                        score,
                        angle,
                        polygon,
                        piece_width,
                        piece_height,
                    )

        if best_new_column is None:
            continue

        _, angle, polygon, piece_width, piece_height = best_new_column
        current_x = new_x
        current_y = margin
        column_width = piece_width

        candidate = affinity.translate(
            polygon,
            xoff=current_x,
            yoff=current_y,
        )
        placements.append(
            Placement(
                reference=piece.reference_display,
                source_name=piece.source_name,
                sheet_index=0,
                rotation=angle,
                polygon=candidate,
                original_polygon=piece.polygon,
                copy_index=copy_index,
                thickness=piece.thickness,
                material=piece.material,
            )
        )
        current_y += piece_height + clearance

    safe: list[Placement] = []
    sheet_inner = box(
        margin,
        margin,
        width - margin,
        height - margin,
    )

    for item in placements:
        if fits_on_sheet(
            item.polygon,
            [placed.polygon for placed in safe],
            sheet_inner,
            clearance,
        ):
            safe.append(item)

    return safe



def fill_one_variable_sheet(
    remaining: list[tuple[DxfPiece, int]],
    width: float,
    height: float,
    margin: float,
    clearance: float,
    rotations: list[int],
    allow_hole_nesting: bool,
    quality_mode: str,
    rotation_cache: dict[tuple[str, int], Polygon],
    deadline: float,
) -> list[Placement]:
    if not remaining or time.perf_counter() >= deadline:
        return []

    attempts, quality_level, max_candidates = quality_parameters(quality_mode)
    best_placements: list[Placement] = []
    best_score = None

    if is_height_priority():
        column_deadline = min(
            deadline,
            time.perf_counter()
            + max(2.0, (deadline - time.perf_counter()) * 0.25),
        )
        column_placements = vertical_column_fill_one_sheet(
            remaining,
            width,
            height,
            margin,
            clearance,
            rotations,
            rotation_cache,
            column_deadline,
        )

        if column_placements:
            used_area = sum(
                item.original_polygon.area
                for item in column_placements
            )
            max_x = max(
                item.polygon.bounds[2]
                for item in column_placements
            )
            max_y = max(
                item.polygon.bounds[3]
                for item in column_placements
            )
            best_placements = column_placements
            best_score = (
                -len(column_placements),
                -round(used_area, 2),
                round(max_x, 2),
                round(max_y, 2),
            )

    for attempt_index in range(attempts):
        if time.perf_counter() >= deadline:
            break

        ordered = order_piece_copies(
            remaining,
            attempt_index,
        )

        placements, _, _, timed_out = nest_one_order(
            ordered,
            width,
            height,
            margin,
            clearance,
            rotations,
            allow_hole_nesting,
            quality_level,
            max_candidates,
            rotation_cache,
            deadline,
            sheet_limit=1,
        )

        if timed_out:
            break

        used_area = sum(
            item.original_polygon.area
            for item in placements
        )

        if placements:
            max_x = max(item.polygon.bounds[2] for item in placements)
            max_y = max(item.polygon.bounds[3] for item in placements)
        else:
            max_x = width
            max_y = height

        if is_height_priority():
            score = (
                -len(placements),
                -round(used_area, 2),
                round(max_x, 2),
                round(max_y, 2),
            )
        else:
            score = (
                -len(placements),
                -round(used_area, 2),
                round(max_y, 2),
                round(max_x, 2),
            )

        if best_score is None or score < best_score:
            best_score = score
            best_placements = placements

    return best_placements


def quick_shelf_pack_expanded(
    remaining: list[tuple[DxfPiece, int]],
    sheet_width: float,
    sheet_height: float,
    margin: float,
    clearance: float,
    rotations: list[int],
    rotation_cache: dict[tuple[str, int], Polygon],
) -> tuple[list[Placement], list[list[Placement]], list[str]]:
    inner_right = sheet_width - margin
    inner_top = sheet_height - margin
    simple_rotations = [
        angle
        for angle in (0, 90, 180, 270)
        if angle in rotations
    ] or [rotations[0] if rotations else 0]

    ordered = sorted(
        remaining,
        key=lambda item: item[0].polygon.area,
        reverse=True,
    )

    sheets_data = []
    unplaced = []

    for piece, copy_index in ordered:
        versions = []

        for angle in simple_rotations:
            cache_key = (piece.reference_key, angle)

            if cache_key not in rotation_cache:
                rotation_cache[cache_key] = rotated_at_origin(
                    piece.polygon,
                    angle,
                )

            polygon = rotation_cache[cache_key]
            width = polygon.bounds[2] - polygon.bounds[0]
            height = polygon.bounds[3] - polygon.bounds[1]
            versions.append((angle, polygon, width, height))

        best = None

        for sheet_index, sheet_data in enumerate(sheets_data):
            for shelf_index, shelf in enumerate(sheet_data["shelves"]):
                for angle, polygon, width, height in versions:
                    if (
                        height <= shelf["height"] + 1e-6
                        and shelf["x"] + width <= inner_right + 1e-6
                    ):
                        score = (
                            0,
                            inner_right - (shelf["x"] + width),
                            shelf["y"],
                        )

                        if best is None or score < best[0]:
                            best = (
                                score,
                                sheet_index,
                                shelf_index,
                                angle,
                                polygon,
                                width,
                                height,
                            )

            new_y = (
                margin
                if not sheet_data["shelves"]
                else max(
                    shelf["y"] + shelf["height"] + clearance
                    for shelf in sheet_data["shelves"]
                )
            )

            for angle, polygon, width, height in versions:
                if (
                    margin + width <= inner_right + 1e-6
                    and new_y + height <= inner_top + 1e-6
                ):
                    score = (1, new_y + height, width)

                    if best is None or score < best[0]:
                        best = (
                            score,
                            sheet_index,
                            None,
                            angle,
                            polygon,
                            width,
                            height,
                        )

        if best is None:
            valid = [
                (height * width, angle, polygon, width, height)
                for angle, polygon, width, height in versions
                if (
                    margin + width <= inner_right + 1e-6
                    and margin + height <= inner_top + 1e-6
                )
            ]

            if not valid:
                unplaced.append(
                    f"{piece.reference_display} - copie {copy_index}"
                )
                continue

            _, angle, polygon, width, height = min(valid)
            sheet_index = len(sheets_data)
            sheets_data.append(
                {
                    "placements": [],
                    "shelves": [
                        {
                            "x": margin,
                            "y": margin,
                            "height": height,
                        }
                    ],
                }
            )
            shelf_index = 0
        else:
            (
                _,
                sheet_index,
                shelf_index,
                angle,
                polygon,
                width,
                height,
            ) = best

            if shelf_index is None:
                sheet_data = sheets_data[sheet_index]
                new_y = (
                    margin
                    if not sheet_data["shelves"]
                    else max(
                        shelf["y"] + shelf["height"] + clearance
                        for shelf in sheet_data["shelves"]
                    )
                )
                sheet_data["shelves"].append(
                    {
                        "x": margin,
                        "y": new_y,
                        "height": height,
                    }
                )
                shelf_index = len(sheet_data["shelves"]) - 1

        shelf = sheets_data[sheet_index]["shelves"][shelf_index]
        placed_polygon = affinity.translate(
            polygon,
            xoff=shelf["x"],
            yoff=shelf["y"],
        )

        placement = Placement(
            reference=piece.reference_display,
            source_name=piece.source_name,
            sheet_index=sheet_index,
            rotation=angle,
            polygon=placed_polygon,
            original_polygon=piece.polygon,
            copy_index=copy_index,
            thickness=piece.thickness,
            material=piece.material,
        )

        sheets_data[sheet_index]["placements"].append(placement)
        shelf["x"] += width + clearance
        shelf["height"] = max(shelf["height"], height)

    sheets = [
        item["placements"]
        for item in sheets_data
    ]
    placements = [
        placement
        for sheet in sheets
        for placement in sheet
    ]

    return placements, sheets, unplaced


def insert_item_variable(
    item: Placement,
    target_sheets: list[list[Placement]],
    target_metas: list[SheetMeta],
    margin: float,
    clearance: float,
    rotations: list[int],
    allow_hole_nesting: bool,
    rotation_cache: dict[tuple[str, int], Polygon],
    deadline: float,
) -> bool:
    reference_key = normalize_reference(item.reference)
    best = None

    for target_index, (target, meta) in enumerate(
        zip(target_sheets, target_metas)
    ):
        if time.perf_counter() >= deadline:
            return False

        fixed = [placed.polygon for placed in target]
        sheet_inner = box(
            margin,
            margin,
            meta.width - margin,
            meta.height - margin,
        )

        current_max_x = max(
            (polygon.bounds[2] for polygon in fixed),
            default=margin,
        )
        current_max_y = max(
            (polygon.bounds[3] for polygon in fixed),
            default=margin,
        )
        current_area = sum(polygon.area for polygon in fixed)

        item_piece = piece_like_from_placement(item)

        for angle in piece_rotation_angles(item_piece, rotations):
            cache_key = (reference_key, angle)

            if cache_key not in rotation_cache:
                rotation_cache[cache_key] = rotated_at_origin(
                    item.original_polygon,
                    angle,
                )

            rotated = rotation_cache[cache_key]
            width = rotated.bounds[2] - rotated.bounds[0]
            height = rotated.bounds[3] - rotated.bounds[1]

            if (
                width > meta.width - 2 * margin + 1e-6
                or height > meta.height - 2 * margin + 1e-6
            ):
                continue

            positions = candidate_positions(
                rotated,
                fixed,
                margin,
                clearance,
                meta.width,
                meta.height,
                allow_hole_nesting,
                2,
                180,
            )

            for x, y in positions:
                candidate = affinity.translate(
                    rotated,
                    xoff=x,
                    yoff=y,
                )

                if not fits_on_sheet(
                    candidate,
                    fixed,
                    sheet_inner,
                    clearance,
                ):
                    continue

                candidate = directional_compact(
                    candidate,
                    fixed,
                    sheet_inner,
                    clearance,
                    minimum_step=1.0,
                )

                score = placement_score(
                    candidate,
                    current_max_x,
                    current_max_y,
                    current_area,
                    fixed,
                    sheet_inner,
                    clearance,
                    margin,
                )

                if best is None or score < best[0]:
                    best = (
                        score,
                        target_index,
                        angle,
                        candidate,
                    )

    if best is None:
        return False

    _, target_index, angle, candidate = best
    target_sheets[target_index].append(
        Placement(
            reference=item.reference,
            source_name=item.source_name,
            sheet_index=target_index,
            rotation=angle,
            polygon=candidate,
            original_polygon=item.original_polygon,
            copy_index=item.copy_index,
            thickness=item.thickness,
            material=item.material,
        )
    )
    return True


def consolidate_variable_sheets(
    sheets: list[list[Placement]],
    metas: list[SheetMeta],
    margin: float,
    clearance: float,
    rotations: list[int],
    allow_hole_nesting: bool,
    deadline: float,
) -> tuple[list[list[Placement]], list[SheetMeta]]:
    rotation_cache: dict[tuple[str, int], Polygon] = {}

    while len(sheets) > 1 and time.perf_counter() < deadline:
        source_order = sorted(
            range(len(sheets)),
            key=lambda index: (
                0 if metas[index].source == "Achat" else 1,
                sum(
                    item.original_polygon.area
                    for item in sheets[index]
                )
                / max(1.0, metas[index].width * metas[index].height),
            ),
        )

        improved = False

        for source_index in source_order:
            if time.perf_counter() >= deadline:
                break

            candidate_sheets = [
                list(sheet)
                for index, sheet in enumerate(sheets)
                if index != source_index
            ]
            candidate_metas = [
                meta
                for index, meta in enumerate(metas)
                if index != source_index
            ]

            source_items = sorted(
                sheets[source_index],
                key=lambda item: item.original_polygon.area,
                reverse=True,
            )

            success = True

            for item in source_items:
                if not insert_item_variable(
                    item,
                    candidate_sheets,
                    candidate_metas,
                    margin,
                    clearance,
                    rotations,
                    allow_hole_nesting,
                    rotation_cache,
                    deadline,
                ):
                    success = False
                    break

            if success:
                sheets = candidate_sheets
                metas = candidate_metas
                improved = True
                break

        if not improved:
            break

    for sheet_index, (sheet, meta) in enumerate(zip(sheets, metas)):
        meta.sheet_index = sheet_index

        for placement in sheet:
            placement.sheet_index = sheet_index

    return sheets, metas


def strategy_score(
    placements: list[Placement],
    metas: list[SheetMeta],
    unplaced: list[str],
    objective: str,
) -> tuple:
    purchase_count = sum(
        1 for meta in metas if meta.source == "Achat"
    )
    total_sheet_area = sum(
        meta.width * meta.height
        for meta in metas
    )
    piece_area = sum(
        item.original_polygon.area
        for item in placements
    )
    waste_area = max(0.0, total_sheet_area - piece_area)

    if objective == "Minimiser le nombre de plaques":
        return (
            len(unplaced),
            len(metas),
            purchase_count,
            round(waste_area, 1),
        )

    if objective == "Minimiser la surface consommée":
        return (
            len(unplaced),
            round(total_sheet_area, 1),
            purchase_count,
            len(metas),
        )

    return (
        len(unplaced),
        purchase_count,
        round(total_sheet_area, 1),
        len(metas),
    )


def optimize_group_with_stock(
    pieces: list[DxfPiece],
    stock_sheets: list[StockSheet],
    purchase_width: float,
    purchase_height: float,
    allow_purchase: bool,
    margin: float,
    clearance: float,
    rotations: list[int],
    allow_hole_nesting: bool,
    quality_mode: str,
    objective: str,
    time_budget_seconds: float,
    progress_callback=None,
) -> tuple[
    list[Placement],
    list[SheetMeta],
    list[str],
    set[str],
]:
    expanded = expand_piece_copies(pieces)
    rotation_cache: dict[tuple[str, int], Polygon] = {}
    start_time = time.perf_counter()
    global_deadline = start_time + max(15.0, float(time_budget_seconds))

    orderings = stock_orderings(stock_sheets, objective)

    if not orderings:
        orderings = [[]]

    if quality_mode == "Rapide":
        orderings = orderings[:1]
    elif quality_mode == "Équilibré":
        orderings = orderings[:2]

    best_result = None

    for strategy_index, ordered_stock in enumerate(orderings):
        if time.perf_counter() >= global_deadline and best_result is not None:
            break

        remaining = list(expanded)
        strategy_sheets: list[list[Placement]] = []
        strategy_metas: list[SheetMeta] = []
        used_stock_ids: set[str] = set()

        sheets_left = max(1, len(ordered_stock))
        strategy_remaining_time = max(
            5.0,
            global_deadline - time.perf_counter(),
        )
        per_sheet_budget = max(
            2.0,
            strategy_remaining_time / (sheets_left + 2),
        )

        for stock_position, stock_sheet in enumerate(ordered_stock):
            if not remaining:
                break

            if time.perf_counter() >= global_deadline:
                break

            local_deadline = min(
                global_deadline,
                time.perf_counter() + per_sheet_budget,
            )

            selected = fill_one_variable_sheet(
                remaining,
                stock_sheet.width,
                stock_sheet.height,
                margin,
                clearance,
                rotations,
                allow_hole_nesting,
                quality_mode,
                rotation_cache,
                local_deadline,
            )

            if not selected:
                continue

            sheet_index = len(strategy_sheets)

            for placement in selected:
                placement.sheet_index = sheet_index

            strategy_sheets.append(selected)
            strategy_metas.append(
                SheetMeta(
                    sheet_index=sheet_index,
                    sheet_id=stock_sheet.instance_id,
                    material=pieces[0].material,
                    thickness=pieces[0].thickness,
                    width=stock_sheet.width,
                    height=stock_sheet.height,
                    source="Stock",
                    stock_type=stock_sheet.stock_type,
                    stock_row_id=stock_sheet.row_id,
                    priority=stock_sheet.priority,
                )
            )
            used_stock_ids.add(stock_sheet.instance_id)
            remaining = remove_placed_copies(
                remaining,
                selected,
            )

            if progress_callback is not None and ordered_stock:
                progress_callback(
                    min(
                        0.65,
                        0.65
                        * (stock_position + 1)
                        / len(ordered_stock),
                    )
                )

        if remaining and allow_purchase:
            purchase_deadline = max(
                global_deadline,
                time.perf_counter() + 8.0,
            )

            attempts, quality_level, max_candidates = quality_parameters(
                quality_mode
            )
            purchase_best = None

            for attempt_index in range(attempts):
                if time.perf_counter() >= purchase_deadline:
                    break

                ordered_remaining = order_piece_copies(
                    remaining,
                    attempt_index,
                )

                placements, sheets, unplaced, timed_out = nest_one_order(
                    ordered_remaining,
                    purchase_width,
                    purchase_height,
                    margin,
                    clearance,
                    rotations,
                    allow_hole_nesting,
                    quality_level,
                    max_candidates,
                    rotation_cache,
                    purchase_deadline,
                )

                if timed_out:
                    break

                score = (
                    len(unplaced),
                    len(sheets),
                    sum(
                        (
                            max(
                                (item.polygon.bounds[2] for item in sheet),
                                default=margin,
                            )
                            * max(
                                (item.polygon.bounds[3] for item in sheet),
                                default=margin,
                            )
                        )
                        for sheet in sheets
                    ),
                )

                if purchase_best is None or score < purchase_best[0]:
                    purchase_best = (
                        score,
                        placements,
                        sheets,
                        unplaced,
                    )

            if purchase_best is None:
                (
                    purchase_placements,
                    purchase_sheets,
                    purchase_unplaced,
                ) = quick_shelf_pack_expanded(
                    remaining,
                    purchase_width,
                    purchase_height,
                    margin,
                    clearance,
                    rotations,
                    rotation_cache,
                )
            else:
                (
                    _,
                    purchase_placements,
                    purchase_sheets,
                    purchase_unplaced,
                ) = purchase_best

            purchase_offset = len(strategy_sheets)

            for local_index, sheet in enumerate(purchase_sheets):
                global_index = purchase_offset + local_index

                for placement in sheet:
                    placement.sheet_index = global_index

                strategy_sheets.append(sheet)
                strategy_metas.append(
                    SheetMeta(
                        sheet_index=global_index,
                        sheet_id=(
                            f"ACHAT-{material_key(pieces[0].material) or 'MAT'}-"
                            f"{thickness_key(pieces[0].thickness)}-"
                            f"{local_index + 1:03d}"
                        ),
                        material=pieces[0].material,
                        thickness=pieces[0].thickness,
                        width=purchase_width,
                        height=purchase_height,
                        source="Achat",
                        stock_type="Tôle neuve",
                        stock_row_id=None,
                        priority=999,
                    )
                )

            remaining = remove_placed_copies(
                remaining,
                purchase_placements,
            )

        unplaced = [
            f"{piece.reference_display} - copie {copy_index}"
            for piece, copy_index in remaining
        ]

        consolidation_deadline = min(
            global_deadline,
            time.perf_counter() + max(
                3.0,
                time_budget_seconds * 0.20,
            ),
        )

        strategy_sheets, strategy_metas = consolidate_variable_sheets(
            strategy_sheets,
            strategy_metas,
            margin,
            clearance,
            rotations,
            allow_hole_nesting,
            consolidation_deadline,
        )

        strategy_placements = [
            item
            for sheet in strategy_sheets
            for item in sheet
        ]

        score = strategy_score(
            strategy_placements,
            strategy_metas,
            unplaced,
            objective,
        )

        result = (
            score,
            strategy_placements,
            strategy_metas,
            unplaced,
            used_stock_ids,
        )

        if best_result is None or score < best_result[0]:
            best_result = result

        if progress_callback is not None:
            progress_callback(
                min(
                    0.98,
                    (strategy_index + 1) / len(orderings),
                )
            )

        if not unplaced and score[1] == 0:
            break

    if progress_callback is not None:
        progress_callback(1.0)

    if best_result is None:
        return [], [], [
            f"{piece.reference_display} - copie {copy_index}"
            for piece, copy_index in expanded
        ], set()

    _, placements, metas, unplaced, used_stock_ids = best_result
    return placements, metas, unplaced, used_stock_ids


def variable_sheet_statistics(
    placements: list[Placement],
    sheet_metas: list[SheetMeta],
) -> pd.DataFrame:
    rows = []

    for meta in sheet_metas:
        sheet_items = [
            item
            for item in placements
            if item.sheet_index == meta.sheet_index
        ]
        used_area = sum(
            item.original_polygon.area
            for item in sheet_items
        )
        sheet_area = meta.width * meta.height
        usage = used_area / sheet_area * 100.0 if sheet_area else 0.0

        rows.append(
            {
                "Tôle": meta.sheet_index + 1,
                "ID plaque": meta.sheet_id,
                "Source": meta.source,
                "Type": meta.stock_type,
                "Matière": meta.material,
                "Épaisseur": meta.thickness,
                "Dimensions (mm)": f"{meta.width:g} × {meta.height:g}",
                "Nombre de pièces": len(sheet_items),
                "Utilisation (%)": round(usage, 2),
                "Chute (%)": round(100.0 - usage, 2),
            }
        )

    return pd.DataFrame(rows)


def variable_placement_table(
    placements: list[Placement],
    sheet_metas: list[SheetMeta],
) -> pd.DataFrame:
    meta_by_index = {
        meta.sheet_index: meta
        for meta in sheet_metas
    }
    rows = []

    for item in placements:
        meta = meta_by_index[item.sheet_index]
        min_x, min_y, max_x, max_y = item.polygon.bounds

        rows.append(
            {
                "Tôle": item.sheet_index + 1,
                "ID plaque": meta.sheet_id,
                "Source": meta.source,
                "Dimensions plaque (mm)": (
                    f"{meta.width:g} × {meta.height:g}"
                ),
                "Repère": item.reference,
                "Copie": item.copy_index,
                "Rotation (°)": item.rotation,
                "X min (mm)": round(min_x, 2),
                "Y min (mm)": round(min_y, 2),
                "Largeur occupée (mm)": round(max_x - min_x, 2),
                "Hauteur occupée (mm)": round(max_y - min_y, 2),
                "Trous": len(item.original_polygon.interiors),
                "Épaisseur": item.thickness,
                "Matière": item.material,
            }
        )

    return pd.DataFrame(rows)


def plot_variable_sheet(
    placements: list[Placement],
    meta: SheetMeta,
):
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.set_xlim(0, meta.width)
    ax.set_ylim(0, meta.height)
    ax.set_aspect("equal", adjustable="box")
    ax.set_title(
        f"{meta.sheet_id} — {meta.width:g} × {meta.height:g} mm "
        f"({meta.source})"
    )
    ax.set_xlabel("X (mm)")
    ax.set_ylabel("Y (mm)")
    ax.grid(True, linewidth=0.3)

    sheet_items = [
        item
        for item in placements
        if item.sheet_index == meta.sheet_index
    ]

    for item in sheet_items:
        x, y = item.polygon.exterior.xy
        filled = ax.fill(x, y, alpha=0.28)
        ax.plot(x, y, linewidth=1)

        for interior in item.polygon.interiors:
            hx, hy = interior.xy
            ax.fill(
                hx,
                hy,
                facecolor=ax.get_facecolor(),
                alpha=1.0,
            )
            ax.plot(hx, hy, linewidth=1)

        point = item.polygon.representative_point()
        ax.text(
            point.x,
            point.y,
            f"{item.reference}\n#{item.copy_index}",
            ha="center",
            va="center",
            fontsize=7,
        )

    return fig


def export_variable_sheets_zip(
    placements: list[Placement],
    sheet_metas: list[SheetMeta],
) -> bytes:
    output = io.BytesIO()

    with zipfile.ZipFile(
        output,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as archive:
        for meta in sheet_metas:
            document = ezdxf.new("R2010")
            modelspace = document.modelspace()

            modelspace.add_lwpolyline(
                [
                    (0, 0),
                    (meta.width, 0),
                    (meta.width, meta.height),
                    (0, meta.height),
                ],
                close=True,
            )

            for item in placements:
                if item.sheet_index == meta.sheet_index:
                    add_polygon_to_dxf(
                        modelspace,
                        item.polygon,
                        f"{item.reference}-{item.copy_index}",
                    )

            safe_sheet_id = re.sub(
                r"[^A-Za-z0-9_-]+",
                "_",
                meta.sheet_id,
            )

            with tempfile.NamedTemporaryFile(
                suffix=".dxf",
                delete=False,
            ) as temporary_file:
                temporary_name = temporary_file.name

            document.saveas(temporary_name)
            archive.write(
                temporary_name,
                arcname=(
                    f"{meta.sheet_index + 1:03d}_"
                    f"{safe_sheet_id}_"
                    f"{meta.width:g}x{meta.height:g}.dxf"
                ),
            )
            Path(temporary_name).unlink(missing_ok=True)

    output.seek(0)
    return output.getvalue()


def stock_balance_dataframe(
    original_stock_dataframe: pd.DataFrame,
    used_sheet_metas: list[SheetMeta],
) -> pd.DataFrame:
    dataframe = ensure_stock_columns(
        original_stock_dataframe
    ).reset_index(drop=True)

    used_by_row = defaultdict(int)

    for meta in used_sheet_metas:
        if (
            meta.source == "Stock"
            and meta.stock_row_id is not None
        ):
            used_by_row[int(meta.stock_row_id)] += 1

    rows = []

    for row_id, row in dataframe.iterrows():
        try:
            initial_quantity = int(
                float(str(row["Quantité"]).replace(",", "."))
            )
        except (TypeError, ValueError):
            initial_quantity = 0

        used_quantity = used_by_row.get(int(row_id), 0)

        rows.append(
            {
                "ID stock": row["ID stock"],
                "Matière": row["Matière"],
                "Épaisseur": row["Épaisseur"],
                "Dimensions (mm)": (
                    f"{row['Largeur (mm)']} × {row['Hauteur (mm)']}"
                ),
                "Type": row["Type"],
                "Quantité initiale": initial_quantity,
                "Quantité utilisée": used_quantity,
                "Quantité restante": max(
                    0,
                    initial_quantity - used_quantity,
                ),
            }
        )

    return pd.DataFrame(rows)



# ============================================================
# Résultats et exports
# ============================================================

def sheet_statistics(
    placements: list[Placement],
    sheet_count: int,
    sheet_width: float,
    sheet_height: float,
) -> pd.DataFrame:
    rows = []

    for sheet_index in range(sheet_count):
        sheet_placements = [
            item for item in placements if item.sheet_index == sheet_index
        ]
        used_area = sum(item.original_polygon.area for item in sheet_placements)
        sheet_area = sheet_width * sheet_height
        usage = used_area / sheet_area * 100 if sheet_area else 0

        first_item = sheet_placements[0] if sheet_placements else None
        rows.append(
            {
                "Tôle": sheet_index + 1,
                "Matière": first_item.material if first_item else "",
                "Épaisseur": first_item.thickness if first_item else "",
                "Nombre de pièces": len(sheet_placements),
                "Surface pièces (mm²)": round(used_area, 1),
                "Utilisation (%)": round(usage, 2),
                "Chute (%)": round(100 - usage, 2),
            }
        )

    return pd.DataFrame(rows)


def placement_table(placements: list[Placement]) -> pd.DataFrame:
    rows = []

    for item in placements:
        min_x, min_y, max_x, max_y = item.polygon.bounds
        rows.append(
            {
                "Tôle": item.sheet_index + 1,
                "Repère": item.reference,
                "Copie": item.copy_index,
                "Rotation (°)": item.rotation,
                "X min (mm)": round(min_x, 2),
                "Y min (mm)": round(min_y, 2),
                "Largeur occupée (mm)": round(max_x - min_x, 2),
                "Hauteur occupée (mm)": round(max_y - min_y, 2),
                "Trous": len(item.original_polygon.interiors),
                "Épaisseur": item.thickness,
                "Matière": item.material,
            }
        )

    return pd.DataFrame(rows)


def plot_sheet(
    placements: list[Placement],
    sheet_index: int,
    sheet_width: float,
    sheet_height: float,
):
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.set_xlim(0, sheet_width)
    ax.set_ylim(0, sheet_height)
    ax.set_aspect("equal", adjustable="box")
    ax.set_title(
        f"Tôle {sheet_index + 1} — {sheet_width:g} × {sheet_height:g} mm"
    )
    ax.set_xlabel("X (mm)")
    ax.set_ylabel("Y (mm)")
    ax.grid(True, linewidth=0.3)

    sheet_items = [
        item for item in placements if item.sheet_index == sheet_index
    ]

    for item in sheet_items:
        x, y = item.polygon.exterior.xy
        filled = ax.fill(x, y, alpha=0.28)
        face_color = filled[0].get_facecolor()
        ax.plot(x, y, linewidth=1)

        # Les trous sont réellement évidés dans l'aperçu.
        for interior in item.polygon.interiors:
            hx, hy = interior.xy
            ax.fill(
                hx,
                hy,
                facecolor=ax.get_facecolor(),
                alpha=1.0,
            )
            ax.plot(hx, hy, linewidth=1)

        point = item.polygon.representative_point()
        ax.text(
            point.x,
            point.y,
            f"{item.reference}\n#{item.copy_index}",
            ha="center",
            va="center",
            fontsize=7,
        )

    return fig


def add_polygon_to_dxf(modelspace, polygon: Polygon, reference: str):
    exterior = [(float(x), float(y)) for x, y in polygon.exterior.coords]
    modelspace.add_lwpolyline(exterior, close=True)

    for interior in polygon.interiors:
        hole = [(float(x), float(y)) for x, y in interior.coords]
        modelspace.add_lwpolyline(hole, close=True)

    point = polygon.representative_point()
    modelspace.add_text(
        reference,
        dxfattribs={"height": 8.0},
    ).set_placement((point.x, point.y))


def export_sheets_zip(
    placements: list[Placement],
    sheet_count: int,
    sheet_width: float,
    sheet_height: float,
) -> bytes:
    output = io.BytesIO()

    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for sheet_index in range(sheet_count):
            document = ezdxf.new("R2010")
            modelspace = document.modelspace()

            modelspace.add_lwpolyline(
                [
                    (0, 0),
                    (sheet_width, 0),
                    (sheet_width, sheet_height),
                    (0, sheet_height),
                ],
                close=True,
            )

            for item in placements:
                if item.sheet_index == sheet_index:
                    add_polygon_to_dxf(
                        modelspace,
                        item.polygon,
                        f"{item.reference}-{item.copy_index}",
                    )

            with tempfile.NamedTemporaryFile(
                suffix=".dxf",
                delete=False,
            ) as tmp:
                temp_name = tmp.name

            document.saveas(temp_name)
            archive.write(
                temp_name,
                arcname=f"tole_{sheet_index + 1}.dxf",
            )
            Path(temp_name).unlink(missing_ok=True)

    output.seek(0)
    return output.getvalue()
# ============================================================
# Interface Streamlit — V13 Stock
# ============================================================

st.set_page_config(
    page_title="OptiTôle Pro v18 — Correctif définitif colonnes",
    page_icon="📐",
    layout="wide",
)

st.title("📐 OptiTôle Pro v18 — Correctif définitif colonnes")
st.caption("Version v18 : correctif définitif max_dimension + remplissage vertical 1500 mm.")

with st.sidebar:
    st.header("Paramètres de découpe")

    purchase_width = st.number_input(
        "Largeur d'une tôle neuve (mm)",
        min_value=100.0,
        value=3000.0,
        step=100.0,
    )

    purchase_height = st.number_input(
        "Hauteur d'une tôle neuve (mm)",
        min_value=100.0,
        value=1500.0,
        step=100.0,
    )

    allow_purchase = st.checkbox(
        "Autoriser l'achat de tôles complémentaires",
        value=True,
        help=(
            "Si le stock disponible ne suffit pas, l'application ajoute "
            "des tôles neuves au format défini ci-dessus."
        ),
    )

    objective = st.selectbox(
        "Objectif principal",
        options=[
            "Minimiser les achats",
            "Minimiser la surface consommée",
            "Minimiser le nombre de plaques",
        ],
        index=0,
    )

    fill_priority_ui = st.selectbox(
        "Sens de remplissage prioritaire",
        options=[
            "Hauteur 1500 mm",
            "Largeur 3000 mm",
            "Automatique dense",
        ],
        index=0,
        help=(
            "Hauteur 1500 mm : l'application remplit d'abord en vertical, "
            "puis avance en largeur. C'est le réglage conseillé pour éviter "
            "une longue bande basse avec du vide au-dessus."
        ),
    )

    margin = st.number_input(
        "Marge extérieure (mm)",
        min_value=0.0,
        value=10.0,
        step=1.0,
    )

    clearance = st.number_input(
        "Espacement entre pièces (mm)",
        min_value=0.0,
        value=5.0,
        step=1.0,
    )

    tolerance = st.number_input(
        "Précision des courbes DXF (mm)",
        min_value=0.1,
        value=1.0,
        step=0.5,
    )

    rotation_step = st.selectbox(
        "Pas de rotation",
        options=[90, 45, 30, 15],
        index=1,
    )
    rotation_options = list(range(0, 360, rotation_step))

    quality_mode = st.selectbox(
        "Qualité d'optimisation",
        options=["Rapide", "Équilibré", "Approfondi", "Maximum", "Pro dense"],
        index=4,
    )

    time_budget_seconds = st.select_slider(
        "Temps maximum global (secondes)",
        options=[90, 180, 300, 600, 900],
        value=300,
    )

    allow_hole_nesting = st.checkbox(
        "Autoriser les petites pièces dans les grands trous",
        value=True,
    )

    natural_edge_angles = st.checkbox(
        "Rotations selon les angles réels des pièces",
        value=True,
        help=(
            "Ajoute automatiquement les orientations des arêtes principales "
            "des pièces. Très utile pour les pièces coudées et inclinées."
        ),
    )

    candidate_multiplier_ui = st.select_slider(
        "Densité de recherche des positions",
        options=[1.0, 1.5, 2.0, 3.0],
        value=2.0,
        help=(
            "Plus la valeur est élevée, plus le moteur essaie de positions. "
            "Le calcul est plus long mais généralement plus dense."
        ),
    )

    st.info(
        "Les plaques de stock compatibles sont essayées en priorité. "
        "Les dimensions réelles de chaque plaque sont respectées."
    )

    st.warning(
        "Le moteur reste heuristique : contrôle obligatoire des DXF "
        "avant toute découpe en production."
    )


st.subheader("1. Stock disponible")

stock_upload_column, stock_download_column = st.columns(2)

with stock_upload_column:
    stock_upload = st.file_uploader(
        "Importer un stock CSV",
        type=["csv"],
        key="stock_csv_upload",
    )

    load_stock_button = st.button(
        "Charger le CSV dans le tableau",
        use_container_width=True,
    )

with stock_download_column:
    template_csv = default_stock_dataframe(
        purchase_width,
        purchase_height,
    ).to_csv(
        index=False,
        sep=";",
    ).encode("utf-8-sig")

    st.download_button(
        "Télécharger un modèle de stock CSV",
        data=template_csv,
        file_name="modele_stock_optitole.csv",
        mime="text/csv",
        use_container_width=True,
    )

if "stock_data" not in st.session_state:
    st.session_state["stock_data"] = default_stock_dataframe(
        purchase_width,
        purchase_height,
    )

if load_stock_button:
    if stock_upload is None:
        st.warning("Sélectionne d'abord un fichier CSV de stock.")
    else:
        try:
            st.session_state["stock_data"] = read_stock_csv(
                stock_upload
            )
            st.session_state.pop("stock_editor", None)
            st.success("Le stock CSV a été chargé.")
            st.rerun()
        except Exception as exc:
            st.error(str(exc))

stock_dataframe = st.data_editor(
    st.session_state["stock_data"],
    num_rows="dynamic",
    use_container_width=True,
    hide_index=True,
    key="stock_editor",
    column_config={
        "Utiliser": st.column_config.CheckboxColumn(
            "Utiliser",
            default=True,
        ),
        "ID stock": st.column_config.TextColumn(
            "ID stock",
            help="Référence interne de la plaque ou de la chute.",
        ),
        "Matière": st.column_config.TextColumn(
            "Matière",
            help="Exemples : S235JR, S355, Inox. Utilise * pour toutes.",
        ),
        "Épaisseur": st.column_config.TextColumn(
            "Épaisseur",
            help="Exemples : 5, 6, 10.",
        ),
        "Largeur (mm)": st.column_config.NumberColumn(
            "Largeur (mm)",
            min_value=1.0,
            format="%.1f",
        ),
        "Hauteur (mm)": st.column_config.NumberColumn(
            "Hauteur (mm)",
            min_value=1.0,
            format="%.1f",
        ),
        "Quantité": st.column_config.NumberColumn(
            "Quantité",
            min_value=0,
            step=1,
            format="%d",
        ),
        "Type": st.column_config.SelectboxColumn(
            "Type",
            options=[
                "Tôle complète",
                "Chute rectangulaire",
                "Réservée",
            ],
        ),
        "Priorité": st.column_config.NumberColumn(
            "Priorité",
            help="1 est prioritaire sur 10.",
            step=1,
            format="%d",
        ),
    },
)

st.session_state["stock_data"] = stock_dataframe

current_stock_csv = ensure_stock_columns(
    stock_dataframe
).to_csv(
    index=False,
    sep=";",
).encode("utf-8-sig")

st.download_button(
    "Télécharger le stock actuellement saisi",
    data=current_stock_csv,
    file_name="stock_disponible_optitole.csv",
    mime="text/csv",
)

st.caption(
    "Chaque ligne peut représenter plusieurs plaques identiques grâce à la "
    "colonne Quantité. Les chutes doivent être rectangulaires dans cette version."
)


st.subheader("2. Charger les fichiers de fabrication")

left, right = st.columns(2)

with left:
    dxf_zip = st.file_uploader(
        "ZIP contenant les fichiers DXF",
        type=["zip"],
    )

with right:
    nomenclature_file = st.file_uploader(
        "Nomenclature Excel ou CSV",
        type=["xlsx", "csv"],
    )


st.subheader("3. Lancer l'analyse et l'optimisation")

run_button = st.button(
    "Analyser le stock, les DXF et optimiser",
    type="primary",
    use_container_width=True,
)

clear_results = st.button(
    "Effacer les anciens résultats",
    use_container_width=True,
)

if clear_results:
    st.session_state.pop("optitole_result", None)
    st.success("Les anciens résultats ont été effacés.")

if run_button:
    if dxf_zip is None or nomenclature_file is None:
        st.warning("Charge le ZIP des DXF et la nomenclature.")
        st.stop()

    try:
        USE_NATURAL_EDGE_ANGLES = bool(natural_edge_angles)
        DENSE_CANDIDATE_MULTIPLIER = float(candidate_multiplier_ui)
        FILL_PRIORITY = str(fill_priority_ui)

        stock_instances = prepare_stock_instances(
            stock_dataframe
        )

        with st.spinner("Lecture de la nomenclature..."):
            nomenclature = read_nomenclature(
                nomenclature_file
            )

        with st.spinner("Lecture et contrôle des DXF..."):
            dxf_files = read_dxf_zip(dxf_zip)
            pieces, missing, unused = build_pieces(
                nomenclature,
                dxf_files,
                tolerance,
            )

        if not pieces:
            st.error("Aucune pièce ne peut être optimisée.")
            st.stop()

        dxf_diagnostic = pd.DataFrame(
            [
                {
                    "Repère": piece.reference_display,
                    "Fichier": piece.source_name,
                    "Trous détectés": len(
                        piece.polygon.interiors
                    ),
                    "Surface nette (mm²)": round(
                        piece.polygon.area,
                        1,
                    ),
                }
                for piece in pieces
            ]
        )

        groups = defaultdict(list)

        for piece in pieces:
            groups[
                (piece.material, piece.thickness)
            ].append(piece)

        group_weights = {
            group_key: group_complexity_weight(group_pieces)
            for group_key, group_pieces in groups.items()
        }
        total_weight = max(
            1.0,
            sum(group_weights.values()),
        )

        available_stock = list(stock_instances)
        all_placements: list[Placement] = []
        all_sheet_metas: list[SheetMeta] = []
        group_reports = []
        unplaced_messages = []
        global_sheet_offset = 0

        for group_number, (
            (material, thickness),
            group_pieces,
        ) in enumerate(groups.items(), start=1):
            group_stock = [
                item
                for item in available_stock
                if stock_matches_group(
                    item,
                    material,
                    thickness,
                )
            ]

            group_time_budget = max(
                15.0,
                float(time_budget_seconds)
                * group_weights[(material, thickness)]
                / total_weight,
            )

            st.write(
                f"Optimisation : **{material} — épaisseur {thickness}** "
                f"avec {len(group_stock)} plaque(s) de stock compatible(s)."
            )

            progress_bar = st.progress(0)

            def update_group_progress(value):
                progress_bar.progress(
                    min(
                        100,
                        max(0, int(value * 100)),
                    )
                )

            (
                placements,
                sheet_metas,
                unplaced,
                used_stock_ids,
            ) = optimize_group_with_stock(
                group_pieces,
                group_stock,
                purchase_width,
                purchase_height,
                allow_purchase,
                margin,
                clearance,
                sorted(rotation_options),
                allow_hole_nesting,
                quality_mode,
                objective,
                group_time_budget,
                progress_callback=update_group_progress,
            )

            progress_bar.empty()

            for placement in placements:
                placement.sheet_index += global_sheet_offset

            for meta in sheet_metas:
                meta.sheet_index += global_sheet_offset

            all_placements.extend(placements)
            all_sheet_metas.extend(sheet_metas)

            available_stock = [
                item
                for item in available_stock
                if item.instance_id not in used_stock_ids
            ]

            piece_area = sum(
                piece.polygon.area * piece.quantity
                for piece in group_pieces
            )
            sheet_area = sum(
                meta.width * meta.height
                for meta in sheet_metas
            )
            yield_percentage = (
                piece_area / sheet_area * 100.0
                if sheet_area > 0
                else 0.0
            )
            stock_used = sum(
                1
                for meta in sheet_metas
                if meta.source == "Stock"
            )
            purchased = sum(
                1
                for meta in sheet_metas
                if meta.source == "Achat"
            )

            group_reports.append(
                {
                    "Matière": material,
                    "Épaisseur": thickness,
                    "Plaques de stock utilisées": stock_used,
                    "Plaques achetées": purchased,
                    "Total plaques": len(sheet_metas),
                    "Surface plaques (mm²)": round(
                        sheet_area,
                        1,
                    ),
                    "Rendement matière (%)": round(
                        yield_percentage,
                        2,
                    ),
                    "Chute estimée (%)": round(
                        100.0 - yield_percentage,
                        2,
                    ),
                    "Pièces placées": len(placements),
                    "Pièces non placées": len(unplaced),
                }
            )

            if unplaced:
                unplaced_messages.append(
                    f"{material} / {thickness} : "
                    + ", ".join(unplaced)
                )

            global_sheet_offset += len(sheet_metas)

        sheet_summary = variable_sheet_statistics(
            all_placements,
            all_sheet_metas,
        )
        details = variable_placement_table(
            all_placements,
            all_sheet_metas,
        )
        stock_balance = stock_balance_dataframe(
            stock_dataframe,
            all_sheet_metas,
        )

        details_csv = details.to_csv(
            index=False,
            sep=";",
        ).encode("utf-8-sig")

        stock_balance_csv = stock_balance.to_csv(
            index=False,
            sep=";",
        ).encode("utf-8-sig")

        export_zip = export_variable_sheets_zip(
            all_placements,
            all_sheet_metas,
        )

        st.session_state["optitole_result"] = {
            "placements": all_placements,
            "sheet_metas": all_sheet_metas,
            "group_reports": group_reports,
            "details": details,
            "details_csv": details_csv,
            "stock_balance": stock_balance,
            "stock_balance_csv": stock_balance_csv,
            "sheet_summary": sheet_summary,
            "export_zip": export_zip,
            "missing": missing,
            "unused": unused,
            "unplaced_messages": unplaced_messages,
            "dxf_diagnostic": dxf_diagnostic,
        }

    except Exception as exc:
        st.exception(exc)


result = st.session_state.get("optitole_result")

if result is not None:
    placements = result["placements"]
    sheet_metas = result["sheet_metas"]
    group_reports = result["group_reports"]
    details = result["details"]
    details_csv = result["details_csv"]
    stock_balance = result["stock_balance"]
    stock_balance_csv = result["stock_balance_csv"]
    sheet_summary = result["sheet_summary"]
    export_zip = result["export_zip"]
    missing = result["missing"]
    unused = result["unused"]
    unplaced_messages = result["unplaced_messages"]
    dxf_diagnostic = result["dxf_diagnostic"]

    if missing:
        st.error(
            "DXF manquants pour les repères : "
            + ", ".join(missing)
        )

    if unused:
        st.warning(
            "DXF présents mais non utilisés : "
            + ", ".join(unused)
        )

    for message in unplaced_messages:
        st.error(
            "Pièces non placées faute de stock ou de dimensions compatibles : "
            + message
        )

    purchased_count = sum(
        1 for meta in sheet_metas if meta.source == "Achat"
    )
    stock_count = sum(
        1 for meta in sheet_metas if meta.source == "Stock"
    )

    st.success(
        f"Optimisation terminée : {len(placements)} pièce(s), "
        f"{stock_count} plaque(s) de stock utilisée(s), "
        f"{purchased_count} plaque(s) achetée(s)."
    )

    st.subheader("4. Diagnostic des contours DXF")
    st.dataframe(
        dxf_diagnostic,
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("5. Résumé par matière et épaisseur")
    st.dataframe(
        pd.DataFrame(group_reports),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("6. Plaques réellement utilisées")
    st.dataframe(
        sheet_summary,
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("7. Stock restant après optimisation")
    st.dataframe(
        stock_balance,
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("8. Résultats détaillés")
    st.dataframe(
        details,
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("9. Aperçu des plaques")

    if sheet_metas:
        selected_sheet_index = st.selectbox(
            "Choisir la plaque à afficher",
            options=list(range(len(sheet_metas))),
            format_func=lambda index: (
                f"{index + 1} — {sheet_metas[index].sheet_id} — "
                f"{sheet_metas[index].width:g} × "
                f"{sheet_metas[index].height:g} mm — "
                f"{sheet_metas[index].source}"
            ),
            key="selected_stock_sheet",
        )

        figure = plot_variable_sheet(
            placements,
            sheet_metas[selected_sheet_index],
        )
        st.pyplot(
            figure,
            clear_figure=True,
        )

    st.subheader("10. Télécharger les résultats")

    download_1, download_2, download_3 = st.columns(3)

    with download_1:
        st.download_button(
            "Rapport de placement CSV",
            data=details_csv,
            file_name="rapport_optitole_v13.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with download_2:
        st.download_button(
            "Stock restant CSV",
            data=stock_balance_csv,
            file_name="stock_restant_optitole.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with download_3:
        st.download_button(
            "Plaques optimisées DXF",
            data=export_zip,
            file_name="optitole_v13_resultats_dxf.zip",
            mime="application/zip",
            use_container_width=True,
        )

else:
    st.info(
        "Saisis le stock, charge les DXF et la nomenclature, "
        "puis lance l'optimisation."
    )

st.divider()
st.caption(
    "OptiTôle Pro v17 — correctif colonnes 1500 mm et stock variable. "
    "Le résultat doit être contrôlé avant toute découpe réelle."
)
